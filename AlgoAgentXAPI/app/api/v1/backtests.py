from __future__ import annotations

import json
import logging
from io import BytesIO
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from math import ceil
from uuid import UUID, uuid4
from typing import Any

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import String, and_, cast, desc, func, or_, select, text, update
from sqlalchemy.orm import load_only
from sqlalchemy.ext.asyncio import AsyncSession

from ...core.dependencies import get_current_user, get_db, get_user_entitlements
from ...db.compat import as_uuid_or_str
from ...db.models import (
    CreditTransaction,
    CreditTransactionType,
    EquityCurve,
    Instrument,
    JobStatus,
    MarketData,
    PerformanceMetric,
    PnLCalendar,
    Strategy,
    Trade,
)
from ...schemas.backtests import BacktestCostPreviewRequest, BacktestRunRequest
from ...services.backtest_service import BacktestError, BacktestService
from ...services.credits.management import CreditManagementService
from ...services.metrics import MetricsCalculator
from ...services.notification_service import NotificationService
from ...services.pricing.backtest_pricing_service import BacktestPricingService
from ...services.billing.credit_cost_service import CreditCostService
from ...services.backtest_advanced_filters import apply_advanced_filters, build_filter_summary
from ...utils.api_response import success_response

router = APIRouter()
logger = logging.getLogger(__name__)


def _to_float(value, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def _to_int(value, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except Exception:
        return default




def _advanced_filters_to_json(value) -> dict | None:
    if value is None or not bool(getattr(value, "enabled", False)):
        return None
    try:
        if hasattr(value, "model_dump"):
            return value.model_dump(mode="json")
        if isinstance(value, dict):
            return value
    except Exception:
        return None
    return None


def _filter_meta_from_impact(impact: dict | None, payload_filters=None) -> dict:
    filters_json = _advanced_filters_to_json(payload_filters)
    if not impact and not filters_json:
        return {
            "advanced_filters": None,
            "filter_summary": None,
            "candles_before_filter": None,
            "candles_after_filter": None,
            "filter_reduction_pct": None,
        }

    return {
        "advanced_filters": filters_json,
        "filter_summary": (impact or {}).get("summary") or (build_filter_summary(payload_filters) if filters_json else None),
        "candles_before_filter": _to_int((impact or {}).get("total_candles_before_filter"), None),
        "candles_after_filter": _to_int((impact or {}).get("total_candles_after_filter"), None),
        "filter_reduction_pct": _to_float((impact or {}).get("filter_reduction_pct"), None),
    }


async def _get_backtest_filter_meta(db: AsyncSession, backtest_id: str) -> dict:
    columns = [
        "advanced_filters",
        "filter_summary",
        "candles_before_filter",
        "candles_after_filter",
        "filter_reduction_pct",
    ]
    available = []
    for column in columns:
        if await _column_exists(db, "performance_metrics", column):
            available.append(column)
    if not available:
        return _filter_meta_from_impact(None, None)

    select_sql = ", ".join(available)
    try:
        result = await db.execute(
            text(f"SELECT {select_sql} FROM performance_metrics WHERE id::text = :backtest_id LIMIT 1"),
            {"backtest_id": str(backtest_id)},
        )
        row = result.mappings().first()
        if not row:
            return _filter_meta_from_impact(None, None)
        data = dict(row)
        advanced_filters = data.get("advanced_filters")
        if isinstance(advanced_filters, str):
            try:
                advanced_filters = json.loads(advanced_filters)
            except Exception:
                advanced_filters = None
        return {
            "advanced_filters": advanced_filters,
            "filter_summary": data.get("filter_summary"),
            "candles_before_filter": _to_int(data.get("candles_before_filter"), None),
            "candles_after_filter": _to_int(data.get("candles_after_filter"), None),
            "filter_reduction_pct": _to_float(data.get("filter_reduction_pct"), None),
        }
    except Exception as exc:
        logger.warning("Unable to load advanced filter metadata for backtest %s: %s", backtest_id, exc)
        return _filter_meta_from_impact(None, None)

def _decimal(value, fallback: Decimal = Decimal("0")) -> Decimal:
    try:
        if value is None:
            return fallback
        return Decimal(str(value))
    except Exception:
        return fallback


def _as_uuid(value: str) -> UUID:
    return UUID(str(value))


def _parse_date_param(value, field_name: str = "date") -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise HTTPException(status_code=422, detail=f"Invalid {field_name}. Use yyyy-mm-dd or dd-mm-yyyy.")


from zoneinfo import ZoneInfo

MARKET_DATA_TZ = ZoneInfo("Asia/Kolkata")


def _ensure_aware_datetime(value):
    """
    Market candle timestamps in DB are stored as IST local time.
    Do NOT treat naive candle timestamps as UTC.
    """
    if value is None:
        return None

    ts = pd.Timestamp(value)

    if ts.tzinfo is None:
        ts = ts.tz_localize(MARKET_DATA_TZ)
    else:
        ts = ts.tz_convert(MARKET_DATA_TZ)

    return ts.to_pydatetime()


async def _table_exists(db: AsyncSession, table_name: str) -> bool:
    try:
        bind = db.get_bind()
        dialect = bind.dialect.name if bind is not None else ""
        if dialect == "sqlite":
            result = await db.execute(
                text("SELECT name FROM sqlite_master WHERE type = 'table' AND name = :name"),
                {"name": table_name},
            )
            return result.scalar() is not None

        result = await db.execute(
            text(
                """
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = current_schema()
                  AND table_name = :table_name
                LIMIT 1
                """
            ),
            {"table_name": table_name},
        )
        return result.scalar() is not None
    except Exception:
        return False




async def _column_exists(db: AsyncSession, table_name: str, column_name: str) -> bool:
    try:
        bind = db.get_bind()
        dialect = bind.dialect.name if bind is not None else ""
        if dialect == "sqlite":
            result = await db.execute(text(f"PRAGMA table_info({table_name})"))
            return any(row[1] == column_name for row in result.fetchall())

        result = await db.execute(
            text(
                """
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = current_schema()
                  AND table_name = :table_name
                  AND column_name = :column_name
                LIMIT 1
                """
            ),
            {"table_name": table_name, "column_name": column_name},
        )
        return result.scalar() is not None
    except Exception:
        return False
async def _column_udt_name(db: AsyncSession, table_name: str, column_name: str) -> str | None:
    try:
        bind = db.get_bind()
        dialect = bind.dialect.name if bind is not None else ""
        if dialect == "sqlite":
            result = await db.execute(text(f"PRAGMA table_info({table_name})"))
            for row in result.fetchall():
                if row[1] == column_name:
                    return str(row[2]).lower()
            return None

        result = await db.execute(
            text(
                """
                SELECT udt_name
                FROM information_schema.columns
                WHERE table_schema = current_schema()
                  AND table_name = :table_name
                  AND column_name = :column_name
                LIMIT 1
                """
            ),
            {"table_name": table_name, "column_name": column_name},
        )
        value = result.scalar()
        return str(value).lower() if value else None
    except Exception:
        return None


async def _table_columns_meta(db: AsyncSession, table_name: str) -> list[dict[str, str]]:
    bind = db.get_bind()
    dialect = bind.dialect.name if bind is not None else ""
    if dialect == "sqlite":
        result = await db.execute(text(f"PRAGMA table_info({table_name})"))
        rows = []
        for row in result.fetchall():
            rows.append({
                "column_name": row[1],
                "is_nullable": "NO" if row[3] else "YES",
                "data_type": str(row[2]).lower(),
                "udt_name": str(row[2]).lower(),
                "column_default": row[4],
            })
        return rows

    result = await db.execute(
        text(
            """
            SELECT
                column_name,
                is_nullable,
                data_type,
                udt_name,
                column_default
            FROM information_schema.columns
            WHERE table_schema = current_schema()
              AND table_name = :table_name
            ORDER BY ordinal_position
            """
        ),
        {"table_name": table_name},
    )
    return [dict(row._mapping) for row in result]


def _default_perf_metric_value(
    column_name: str,
    payload: BacktestRunRequest,
    service_response,
    metrics: dict,
    winning_trades: int,
    losing_trades: int,
    total_trades: int,
):
    final_capital = _to_float(service_response.final_capital, 0.0)
    initial_capital = _to_float(payload.capital, 0.0)
    net_profit = _to_float(metrics.get("net_profit"), 0.0)
    trade_pnls = [_to_float(getattr(trade, "pnl", 0.0), 0.0) for trade in (service_response.result.trades or [])]
    wins = [p for p in trade_pnls if p > 0]
    losses = [abs(p) for p in trade_pnls if p < 0]
    return_pct = ((final_capital - initial_capital) / initial_capital * 100.0) if initial_capital else 0.0
    expectancy = (net_profit / total_trades) if total_trades else 0.0

    special_values = {
        "id": None,
        "user_id": as_uuid_or_str(payload.user_id) if hasattr(payload, "user_id") else None,
        "period": f"{payload.start_date.isoformat()} to {payload.end_date.isoformat()}",
        "strategy_id": payload.strategy_id,
        "instrument_id": payload.instrument_id,
        "timeframe": payload.timeframe,
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "initial_capital": _decimal(initial_capital),
        "final_capital": _decimal(final_capital),
        "net_profit": _decimal(net_profit),
        "max_drawdown": _decimal(metrics.get("max_drawdown", 0.0)),
        "sharpe_ratio": _decimal(metrics.get("sharpe_ratio", 0.0)),
        "sortino_ratio": _decimal(metrics.get("sortino_ratio", 0.0)),
        "calmar_ratio": _decimal(metrics.get("calmar_ratio", 0.0)),
        "win_rate": _decimal(metrics.get("win_rate", 0.0)),
        "total_trades": total_trades,
        "winning_trades": winning_trades,
        "losing_trades": losing_trades,
        "profit_factor": _decimal(metrics.get("profit_factor", 0.0)),
        "avg_win": _decimal((sum(wins) / len(wins)) if wins else 0.0),
        "avg_loss": _decimal((sum(losses) / len(losses)) if losses else 0.0),
        "expectancy": _decimal(expectancy),
        "return_pct": _decimal(return_pct),
        "status": "completed",
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
    }
    return special_values.get(column_name)


async def _resolve_backtest_fk_value(db: AsyncSession, table_name: str, backtest_id: str):
    udt_name = await _column_udt_name(db, table_name, "backtest_id")
    if udt_name in {"uuid"}:
        return _as_uuid(backtest_id)
    return backtest_id


def _uuid_or_bigint_for_column(meta: dict):
    """Return a safe primary-key value for legacy trades schemas.

    Some AlgoAgentX databases were created with trades.id as BIGINT while other
    migrations used UUID/TEXT.  The report must not lose the entire trade list
    because one legacy id type differs, so generate the id based on the real DB
    column type instead of the SQLAlchemy model assumption.
    """
    data_type = str(meta.get("data_type") or meta.get("udt_name") or "").lower()
    udt_name = str(meta.get("udt_name") or "").lower()
    if udt_name == "uuid" or data_type == "uuid":
        return uuid4()
    if any(token in data_type for token in ["bigint", "integer", "smallint", "int"]):
        return int(uuid4().int % 9_000_000_000_000_000_000)
    return str(uuid4())


def _default_trade_value(
    column_name: str,
    meta: dict,
    *,
    backtest_id: str,
    trades_fk_value,
    user_id: str,
    payload: BacktestRunRequest,
    trade,
):
    """Safe fallback for NOT NULL legacy columns on the trades table."""
    name = str(column_name).lower()
    udt_name = str(meta.get("udt_name") or "").lower()
    data_type = str(meta.get("data_type") or udt_name or "").lower()

    if name == "id":
        return _uuid_or_bigint_for_column(meta)
    if name == "backtest_id":
        return trades_fk_value
    if name == "user_id":
        return _as_uuid(user_id) if udt_name == "uuid" else str(user_id)
    if name == "strategy_id":
        return _as_uuid(payload.strategy_id) if udt_name == "uuid" else str(payload.strategy_id)
    if name == "instrument_id":
        return payload.instrument_id
    if name in {"created_at", "updated_at"} or "time" in data_type:
        if name in {"entry_time", "entry_datetime", "opened_at", "open_time"}:
            return _ensure_aware_datetime(getattr(trade, "entry_datetime", None))
        if name in {"exit_time", "exit_datetime", "closed_at", "close_time"}:
            return _ensure_aware_datetime(getattr(trade, "exit_datetime", None))
        return datetime.utcnow()
    if name == "side":
        return str(getattr(trade, "direction", "") or "")
    if name in {"direction", "trade_side"}:
        return str(getattr(trade, "direction", "") or "")
    if name == "symbol":
        return ""
    if name == "timeframe":
        return payload.timeframe
    if name in {"qty", "quantity", "volume", "lots", "lot_size"}:
        return int(_to_float(getattr(trade, "quantity", 0), 0))
    if name in {"entry_price", "open_price"}:
        return _decimal(getattr(trade, "entry_price", 0))
    if name in {"exit_price", "close_price"}:
        return _decimal(getattr(trade, "exit_price", 0))
    if name in {"pnl", "profit", "net_pnl"}:
        return _decimal(getattr(trade, "pnl", 0))
    if name in {"exit_type", "exit_reason", "status"}:
        return str(getattr(trade, "exit_reason", "closed") or "closed")
    if any(token in data_type for token in ["numeric", "decimal", "double", "real", "float"]):
        return Decimal("0")
    if any(token in data_type for token in ["bigint", "integer", "smallint", "int"]):
        return 0
    if data_type == "date":
        return payload.start_date
    if "bool" in data_type:
        return False
    if udt_name == "uuid":
        return uuid4()
    return ""


def _build_trade_insert_values(
    trade_columns_meta: list[dict],
    *,
    base_values: dict,
    backtest_id: str,
    trades_fk_value,
    user_id: str,
    payload: BacktestRunRequest,
    trade,
) -> dict:
    """Build a DB-compatible trade insert row.

    This keeps Phase 1 compatible with old AlgoAgentX DBs that have extra
    NOT NULL columns on trades. Without this, the insert fails silently inside
    the nested transaction and the report shows 0 trades while equity/PnL exist.
    """
    final_values = {}
    meta_by_column = {meta["column_name"]: meta for meta in trade_columns_meta}
    for column, meta in meta_by_column.items():
        if column in base_values:
            final_values[column] = base_values[column]
            continue
        if str(meta.get("is_nullable", "YES")).upper() == "NO" and not meta.get("column_default"):
            fallback = _default_trade_value(
                column,
                meta,
                backtest_id=backtest_id,
                trades_fk_value=trades_fk_value,
                user_id=user_id,
                payload=payload,
                trade=trade,
            )
            if fallback is not None:
                final_values[column] = fallback
    return final_values


def _trade_transparency_values(trade) -> dict:
    entry_price = _to_float(getattr(trade, "entry_price", None), 0.0)
    exit_price = _to_float(getattr(trade, "exit_price", None), 0.0)
    stop_loss = _to_float(getattr(trade, "stop_loss", None), 0.0)
    target = _to_float(getattr(trade, "target", None), 0.0)
    quantity = _to_float(getattr(trade, "quantity", None), 0.0)
    pnl = _to_float(getattr(trade, "pnl", None), 0.0)
    risk_points = _to_float(getattr(trade, "risk_points", None), 0.0) or abs(entry_price - stop_loss)
    reward_points = _to_float(getattr(trade, "reward_points", None), 0.0) or abs(target - entry_price)
    rr_ratio = _to_float(getattr(trade, "rr_ratio", None), 0.0) or ((reward_points / risk_points) if risk_points > 0 else 0.0)
    risk_amount = _to_float(getattr(trade, "risk_amount", None), 0.0) or (risk_points * quantity)
    reward_amount = _to_float(getattr(trade, "reward_amount", None), 0.0) or (reward_points * quantity)
    r_multiple = _to_float(getattr(trade, "r_multiple", None), 0.0) or ((pnl / risk_amount) if risk_amount > 0 else 0.0)
    return {
        "stop_loss": stop_loss or None,
        "target": target or None,
        "risk_points": risk_points or None,
        "reward_points": reward_points or None,
        "rr_ratio": rr_ratio or None,
        "risk_amount": risk_amount or None,
        "reward_amount": reward_amount or None,
        "r_multiple": r_multiple,
        "signal_reason": getattr(trade, "signal_reason", None),
    }



def _trade_json_from_service(service_response) -> list[dict]:
    """Return complete trade rows as JSON-safe data for report fallback.

    Some deployed databases still have legacy constraints on the trades table.
    Even when the dedicated trades insert fails, the report must not lose trade
    visibility, so we also persist a compact copy in performance_metrics.trade_details.
    """
    rows: list[dict] = []
    for trade in (getattr(getattr(service_response, "result", None), "trades", None) or []):
        risk_values = _trade_transparency_values(trade)
        row = {
            "entry_time": getattr(trade, "entry_datetime", None).isoformat() if getattr(trade, "entry_datetime", None) else None,
            "exit_time": getattr(trade, "exit_datetime", None).isoformat() if getattr(trade, "exit_datetime", None) else None,
            "side": getattr(trade, "direction", None),
            "quantity": _to_float(getattr(trade, "quantity", 0), None),
            "lot_size": _to_float(getattr(trade, "lot_size", None), None),
            "entry_price": _to_float(getattr(trade, "entry_price", None), None),
            "exit_price": _to_float(getattr(trade, "exit_price", None), None),
            "pnl": _to_float(getattr(trade, "pnl", None), None),
            "exit_type": getattr(trade, "exit_reason", None),
            "exit_reason": getattr(trade, "exit_reason", None),
            "account_currency": getattr(trade, "account_currency", None),
            "currency_symbol": getattr(trade, "currency_symbol", None),
            "asset_class": getattr(trade, "asset_class", None),
            "quantity_mode": getattr(trade, "quantity_mode", None),
            "actual_risk_amount": _to_float(getattr(trade, "actual_risk_amount", None), None),
            "risk_ticks": _to_float(getattr(trade, "risk_ticks", None), None),
            "risk_pips": _to_float(getattr(trade, "risk_pips", None), None),
            "reward_ticks": _to_float(getattr(trade, "reward_ticks", None), None),
            "expected_reward_amount": _to_float(getattr(trade, "expected_reward_amount", None), None),
            "sl_mode": getattr(trade, "sl_mode", None),
            "position_size_mode": getattr(trade, "position_size_mode", None),
            "runtime_config_snapshot": getattr(trade, "runtime_config_snapshot", None),
            "instrument_spec_snapshot": getattr(trade, "instrument_spec_snapshot", None),
            "lifecycle_events": getattr(trade, "lifecycle_events", None) or [],
            **{key: (_to_float(value, None) if key != "signal_reason" else value) for key, value in risk_values.items()},
        }
        rows.append(row)
    return rows


def _normalise_trade_detail_rows(value) -> list[dict]:
    """Decode trade_details JSONB/text into report-compatible rows."""
    if value is None:
        return []
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except Exception:
            return []
    if not isinstance(value, list):
        return []
    rows: list[dict] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row["quantity"] = _to_float(row.get("quantity"), None)
        for key in ["entry_price", "exit_price", "stop_loss", "target", "risk_points", "risk_ticks", "risk_pips", "reward_points", "reward_ticks", "rr_ratio", "risk_amount", "actual_risk_amount", "reward_amount", "expected_reward_amount", "r_multiple", "pnl", "lot_size"]:
            row[key] = _to_float(row.get(key), None)
        row["lifecycle_events"] = _jsonish(row.get("lifecycle_events")) or []
        if not row.get("risk_points") or not row.get("reward_points") or not row.get("risk_amount"):
            row.update(_risk_values_from_row(row))
        rows.append(row)
    return rows

def _risk_values_from_row(row: dict) -> dict:
    entry_price = _to_float(row.get("entry_price"), 0.0)
    stop_loss = _to_float(row.get("stop_loss"), 0.0)
    target = _to_float(row.get("target"), 0.0)
    quantity = _to_float(row.get("quantity"), 0.0)
    pnl = _to_float(row.get("pnl"), 0.0)
    risk_points = _to_float(row.get("risk_points"), 0.0) or (abs(entry_price - stop_loss) if stop_loss else 0.0)
    reward_points = _to_float(row.get("reward_points"), 0.0) or (abs(target - entry_price) if target else 0.0)
    rr_ratio = _to_float(row.get("rr_ratio"), 0.0) or ((reward_points / risk_points) if risk_points > 0 else 0.0)
    risk_amount = _to_float(row.get("risk_amount"), 0.0) or (risk_points * quantity)
    reward_amount = _to_float(row.get("reward_amount"), 0.0) or (reward_points * quantity)
    r_multiple = _to_float(row.get("r_multiple"), 0.0) or ((pnl / risk_amount) if risk_amount > 0 else 0.0)
    return {
        "stop_loss": _to_float(row.get("stop_loss"), None),
        "target": _to_float(row.get("target"), None),
        "risk_points": risk_points if risk_points else None,
        "reward_points": reward_points if reward_points else None,
        "rr_ratio": rr_ratio if rr_ratio else None,
        "risk_amount": risk_amount if risk_amount else None,
        "reward_amount": reward_amount if reward_amount else None,
        "r_multiple": r_multiple if risk_amount else None,
        "signal_reason": row.get("signal_reason"),
    }


def _trade_df_from_service(service_response) -> pd.DataFrame:
    if not service_response.result.trades:
        return pd.DataFrame(columns=["entry_time", "exit_time", "pnl"])
    return pd.DataFrame(
        [
            {
                "entry_time": trade.entry_datetime,
                "exit_time": trade.exit_datetime,
                "pnl": _to_float(trade.pnl),
            }
            for trade in service_response.result.trades
        ]
    )


def _equity_df_from_service(service_response, start_date: date) -> pd.DataFrame:
    if not service_response.result.equity_curve:
        return pd.DataFrame(columns=["timestamp", "equity"])

    base = datetime.combine(start_date, time(0, 0, 0))
    rows = []
    for i, value in enumerate(service_response.result.equity_curve):
        rows.append({"timestamp": base + timedelta(minutes=i), "equity": _to_float(value)})
    return pd.DataFrame(rows)


async def _quote_backtest_cost(
    db: AsyncSession,
    plan_code: str | None,
    strategy_id: str | None,
    instrument_id: int | None,
    timeframe: str,
    start_date: date,
    end_date: date,
    use_actual_candle_count: bool,
    candle_count_override: int | None = None,
    candle_count_mode_override: str | None = None,
    advanced_filters=None,
) -> dict:
    """PAY-BILL-5: admin-managed credit expense rules.

    Keep the legacy BacktestPricingService import untouched for compatibility, but
    route backtest debit/preview amounts through the new rule engine.
    """
    if candle_count_override is not None:
        candle_count = int(candle_count_override or 0)
    else:
        try:
            quote = await BacktestPricingService.quote_backtest_cost(
                db,
                timeframe=timeframe,
                start_date=start_date,
                end_date=end_date,
                instrument_id=instrument_id,
                strategy_parameters=None,
                use_actual_candle_count=use_actual_candle_count,
                plan_code=plan_code,
                candle_count_override=None,
                candle_count_mode_override=candle_count_mode_override,
            )
            candle_count = _to_int((quote.get("breakdown") or {}).get("candle_count"), 0)
        except Exception:
            candle_count = 0

    estimate = await CreditCostService.calculate_backtest_credit_cost(
        db,
        user_id=None,
        instrument_id=instrument_id,
        timeframe=timeframe,
        start_date=start_date,
        end_date=end_date,
        candle_count=candle_count,
        advanced_filters=advanced_filters,
    )
    estimate["breakdown"]["candle_count_mode"] = candle_count_mode_override or ("actual" if use_actual_candle_count else "estimated")
    return estimate




async def _advanced_filter_preview(
    db: AsyncSession,
    *,
    instrument_id: int | None,
    timeframe: str,
    start_date: date,
    end_date: date,
    advanced_filters,
) -> dict | None:
    """Calculate AF-2 preview details using the same filter path as actual runs."""
    if instrument_id is None or advanced_filters is None or not bool(getattr(advanced_filters, "enabled", False)):
        return None

    market_data_df = await BacktestService._fetch_market_data(db, instrument_id, timeframe, start_date, end_date)
    instrument_symbol, instrument_market = await BacktestService._get_instrument_details(db, instrument_id)
    _, impact = apply_advanced_filters(
        market_data_df,
        advanced_filters,
        timeframe=timeframe,
        instrument_symbol=instrument_symbol,
        instrument_market=instrument_market,
    )

    filters = impact.get("filters") or {}
    before_count = _to_int(impact.get("total_candles_before_filter"), 0)
    after_count = _to_int(impact.get("total_candles_after_filter"), before_count)
    removed = _to_int(impact.get("candles_removed"), max(before_count - after_count, 0))
    warnings = list(impact.get("warnings") or [])

    return {
        "enabled": True,
        "days_of_week": filters.get("days_of_week") or [],
        "session": filters.get("session") or "ALL",
        "custom_start_time": filters.get("custom_start_time"),
        "custom_end_time": filters.get("custom_end_time"),
        "timezone": filters.get("timezone") or "Asia/Kolkata",
        "summary": build_filter_summary(advanced_filters),
        "status": impact.get("status") or "ok",
        "total_candles_before_filter": before_count,
        "total_candles_after_filter": after_count,
        "candles_removed": removed,
        "filter_reduction_pct": float(impact.get("filter_reduction_pct") or 0.0),
        "minimum_candles_required": _to_int(impact.get("minimum_candles_required"), 0),
        "warning": impact.get("warning"),
        "warnings": warnings,
        "raw_impact": impact,
    }


async def _get_market_data_availability_guard(
    db: AsyncSession,
    instrument_id: int,
    timeframe: str,
    start_date: date,
    end_date: date,
) -> dict:
    requested_start = datetime.combine(start_date, time.min)
    requested_end = datetime.combine(end_date, time.max)

    instrument = await db.get(Instrument, instrument_id)
    symbol = str(getattr(instrument, "symbol", "") or "").upper()
    market = str(getattr(instrument, "market", "") or "").upper()
    is_forex_like = market == "FOREX" or any(token in symbol for token in ["XAU", "XAG", "EUR", "GBP", "JPY", "USD"])
    boundary_tolerance_days = 3 if is_forex_like else 1

    overall = (
        await db.execute(
            select(
                func.min(MarketData.timestamp).label("available_start"),
                func.max(MarketData.timestamp).label("available_end"),
                func.count().label("total_count"),
            ).where(
                MarketData.instrument_id == instrument_id,
                MarketData.timeframe == timeframe,
            )
        )
    ).first()

    ranged = (
        await db.execute(
            select(
                func.min(MarketData.timestamp).label("range_start"),
                func.max(MarketData.timestamp).label("range_end"),
                func.count().label("record_count"),
            ).where(
                MarketData.instrument_id == instrument_id,
                MarketData.timeframe == timeframe,
                MarketData.timestamp >= requested_start,
                MarketData.timestamp <= requested_end,
            )
        )
    ).first()

    available_start = getattr(overall, "available_start", None) if overall else None
    available_end = getattr(overall, "available_end", None) if overall else None
    total_count = int(getattr(overall, "total_count", 0) or 0) if overall else 0
    range_start = getattr(ranged, "range_start", None) if ranged else None
    range_end = getattr(ranged, "range_end", None) if ranged else None
    record_count = int(getattr(ranged, "record_count", 0) or 0) if ranged else 0

    range_start_date = range_start.date() if range_start else None
    range_end_date = range_end.date() if range_end else None
    dataset_start_date = available_start.date() if available_start else None
    dataset_end_date = available_end.date() if available_end else None

    missing_before = False
    missing_after = False
    if record_count > 0 and range_start_date:
        missing_before = (range_start_date - start_date).days > boundary_tolerance_days
    elif dataset_start_date:
        missing_before = (dataset_start_date - start_date).days > boundary_tolerance_days

    if record_count > 0 and range_end_date:
        missing_after = (end_date - range_end_date).days > boundary_tolerance_days
    elif dataset_end_date:
        missing_after = (end_date - dataset_end_date).days > boundary_tolerance_days

    dataset_missing = total_count <= 0 or available_start is None or available_end is None
    blocked = dataset_missing or record_count <= 0 or missing_before or missing_after

    if dataset_missing:
        message = "Market data is missing for this instrument/timeframe/date range. Ask admin to import missing candles."
        status_value = "error"
    elif record_count <= 0:
        message = "Market data is missing for this instrument/timeframe/date range. Ask admin to import missing candles."
        status_value = "error"
    elif blocked:
        message = "Market data is missing for this instrument/timeframe/date range. Ask admin to import missing candles."
        status_value = "error"
    else:
        message = "Market data available for selected range."
        status_value = "ok"

    return {
        "status": status_value,
        "available": not blocked,
        "blocked": blocked,
        "message": message,
        "instrument_id": instrument_id,
        "timeframe": timeframe,
        "requested_start": requested_start,
        "requested_end": requested_end,
        "requested_start_iso": requested_start.isoformat(),
        "requested_end_iso": requested_end.isoformat(),
        "available_start": range_start,
        "available_end": range_end,
        "available_start_iso": range_start.isoformat() if range_start else None,
        "available_end_iso": range_end.isoformat() if range_end else None,
        "dataset_start_iso": available_start.isoformat() if available_start else None,
        "dataset_end_iso": available_end.isoformat() if available_end else None,
        "missing_before": missing_before,
        "missing_after": missing_after,
        "record_count": record_count,
        "total_count": total_count,
        "boundary_tolerance_days": boundary_tolerance_days,
        "is_forex_like": is_forex_like,
    }


def _raise_market_data_unavailable(availability: dict) -> None:
    raise HTTPException(
        status_code=400,
        detail={
            "code": "MARKET_DATA_UNAVAILABLE",
            "message": availability.get("message") or (
                "Market data is missing for this instrument/timeframe/date range. "
                "Ask admin to import missing candles."
            ),
            "available_start": availability.get("available_start_iso"),
            "available_end": availability.get("available_end_iso"),
            "dataset_start": availability.get("dataset_start_iso"),
            "dataset_end": availability.get("dataset_end_iso"),
            "requested_start": availability.get("requested_start_iso"),
            "requested_end": availability.get("requested_end_iso"),
            "missing_before": bool(availability.get("missing_before")),
            "missing_after": bool(availability.get("missing_after")),
            "record_count": int(availability.get("record_count") or 0),
            "action": "Ask admin to import missing candles for this instrument/timeframe/date range.",
        },
    )


async def _get_backtest_debit_map(db: AsyncSession, backtest_ids: list[str]) -> dict[str, dict]:
    if not backtest_ids:
        return {}

    rows = (
        await db.execute(
            select(
                CreditTransaction.backtest_id,
                CreditTransaction.id,
                CreditTransaction.amount,
                CreditTransaction.transaction_type,
                CreditTransaction.source,
            )
            .where(CreditTransaction.backtest_id.in_(backtest_ids))
            .order_by(CreditTransaction.created_at.desc())
        )
    ).all()

    debit_map: dict[str, dict] = {}
    for backtest_id, txn_id, amount, transaction_type, source in rows:
        key = str(backtest_id)
        if key not in debit_map:
            debit_map[key] = {
                "debit_transaction_id": None,
                "credit_cost": 0.0,
                "included_debited": 0.0,
                "wallet_debited": 0.0,
                "included_refunded": 0.0,
                "wallet_refunded": 0.0,
                "effective_credit_cost": 0.0,
            }

        tx_type = transaction_type.name if hasattr(transaction_type, "name") else str(transaction_type).upper()
        tx_source = str(source or "").lower()
        tx_amount = _to_float(amount)

        if tx_type == CreditTransactionType.DEBIT.name:
            if debit_map[key]["debit_transaction_id"] is None:
                debit_map[key]["debit_transaction_id"] = str(txn_id)

            debit_map[key]["credit_cost"] += tx_amount
            if tx_source == CreditManagementService.SOURCE_BACKTEST_INCLUDED_DEBIT:
                debit_map[key]["included_debited"] += tx_amount
            else:
                debit_map[key]["wallet_debited"] += tx_amount

        if tx_type == CreditTransactionType.REFUND.name:
            if tx_source == CreditManagementService.SOURCE_BACKTEST_INCLUDED_REFUND:
                debit_map[key]["included_refunded"] += tx_amount
            else:
                debit_map[key]["wallet_refunded"] += tx_amount

    for key in debit_map.keys():
        debit_total = float(debit_map[key]["credit_cost"])
        refund_total = float(debit_map[key]["included_refunded"] + debit_map[key]["wallet_refunded"])
        debit_map[key]["effective_credit_cost"] = max(debit_total - refund_total, 0.0)

    return debit_map


async def _serialize_summary(
    db: AsyncSession,
    row: PerformanceMetric,
    strategy_name: str | None = None,
    instrument_symbol: str | None = None,
    credit_cost: float | None = None,
    debit_transaction_id: str | None = None,
) -> dict:
    if strategy_name is None and row.strategy_id:
        strategy_name = (
            await db.execute(select(Strategy.name).where(Strategy.id == row.strategy_id))
        ).scalar_one_or_none()
    instrument_meta = None
    if row.instrument_id:
        instrument_meta = (
            await db.execute(
                select(Instrument).where(Instrument.id == row.instrument_id)
            )
        ).scalars().first()
        if instrument_symbol is None and instrument_meta is not None:
            instrument_symbol = instrument_meta.symbol

    filter_meta = await _get_backtest_filter_meta(db, str(row.id))

    overlay = await _professional_summary_overlay(db, str(row.id), [])
    instrument_spec = overlay.get("instrument_spec_snapshot") if isinstance(overlay.get("instrument_spec_snapshot"), dict) else {}
    if instrument_meta is not None:
        instrument_spec = {
            **(instrument_spec or {}),
            "symbol": getattr(instrument_meta, "symbol", None),
            "asset_class": getattr(instrument_meta, "asset_class", None),
            "account_currency": getattr(instrument_meta, "account_currency", None),
            "currency_symbol": getattr(instrument_meta, "currency_symbol", None),
            "quantity_mode": getattr(instrument_meta, "quantity_mode", None),
        }
    currency_payload = _infer_currency_payload(
        instrument_symbol=instrument_symbol,
        asset_class=overlay.get("asset_class") or (instrument_spec or {}).get("asset_class"),
        account_currency=overlay.get("account_currency") or (instrument_spec or {}).get("account_currency"),
        currency_symbol=overlay.get("currency_symbol") or (instrument_spec or {}).get("currency_symbol"),
        quantity_mode=overlay.get("quantity_mode") or (instrument_spec or {}).get("quantity_mode"),
        instrument_spec=instrument_spec,
    )

    return {
        "id": str(row.id),
        "strategy_id": row.strategy_id,
        "strategy_name": strategy_name,
        "instrument_id": row.instrument_id,
        "instrument_symbol": instrument_symbol,
        "timeframe": row.timeframe,
        "start_date": row.start_date.isoformat() if row.start_date else None,
        "end_date": row.end_date.isoformat() if row.end_date else None,
        "initial_capital": _to_float(row.initial_capital),
        "final_capital": _to_float(row.final_capital),
        "net_profit": _to_float(row.net_profit),
        "max_drawdown": _to_float(row.max_drawdown),
        "sharpe_ratio": _to_float(row.sharpe_ratio),
        "win_rate": _to_float(row.win_rate),
        "total_trades": _to_int(row.total_trades),
        "winning_trades": _to_int(getattr(row, "winning_trades", 0)),
        "losing_trades": _to_int(getattr(row, "losing_trades", 0)),
        "profit_factor": _to_float(getattr(row, "profit_factor", None)),
        "credit_cost": credit_cost,
        "effective_credit_cost": credit_cost,
        "debit_transaction_id": debit_transaction_id,
        "status": row.status,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "updated_at": (row.__dict__.get("updated_at").isoformat() if row.__dict__.get("updated_at") else None),
        "return_pct": _to_float(getattr(row, "return_pct", None), None),
        "avg_win": _to_float(getattr(row, "avg_win", None), None),
        "avg_loss": _to_float(getattr(row, "avg_loss", None), None),
        "expectancy": _to_float(getattr(row, "expectancy", None), None),
        "account_currency": currency_payload.get("account_currency"),
        "currency_symbol": currency_payload.get("currency_symbol"),
        "asset_class": currency_payload.get("asset_class"),
        "quantity_mode": currency_payload.get("quantity_mode"),
        "is_legacy_currency": currency_payload.get("is_legacy_currency"),
        "position_size_mode": overlay.get("position_size_mode"),
        "sl_mode": overlay.get("sl_mode"),
        "rr_ratio": _to_float(overlay.get("rr_ratio"), None),
        "risk_percent": _to_float(overlay.get("risk_percent"), None),
        "avg_lot_size": _to_float(overlay.get("avg_lot_size"), None),
        "avg_quantity": _to_float(overlay.get("avg_quantity"), None),
        # asyncpg expects JSON/JSONB values passed through raw text() SQL to be
        # JSON-encoded strings. Passing a Python dict here raises:
        #   dict object has no attribute encode
        # Keep reads backward-compatible by decoding strings in _get_backtest_filter_meta().
        "advanced_filters": json.dumps(filter_meta.get("advanced_filters")) if filter_meta.get("advanced_filters") is not None else None,
        "filter_summary": filter_meta.get("filter_summary"),
        "candles_before_filter": filter_meta.get("candles_before_filter"),
        "candles_after_filter": filter_meta.get("candles_after_filter"),
        "filter_reduction_pct": filter_meta.get("filter_reduction_pct"),
        "runtime_config_snapshot": overlay.get("runtime_config_snapshot"),
        "instrument_spec_snapshot": overlay.get("instrument_spec_snapshot"),
        "runtime_summary": _runtime_summary_text(overlay.get("runtime_config_snapshot"), filter_meta.get("filter_summary")),
    }


async def _save_backtest_payload(
    db: AsyncSession,
    *,
    user_id: str,
    payload: BacktestRunRequest,
    service_response,
    metrics: dict,
) -> str:
    backtest_uuid = uuid4()
    backtest_id = str(backtest_uuid)

    win_rate = Decimal(str(metrics.get("win_rate", 0.0)))
    total_trades = int(service_response.total_trades or 0)
    winning_trades = int(round(total_trades * float(win_rate))) if total_trades > 0 else 0
    losing_trades = max(total_trades - winning_trades, 0)

    filter_meta = _filter_meta_from_impact(service_response.advanced_filter_impact, payload.advanced_filters)

    insert_values = {
        "id": backtest_id,
        "user_id": as_uuid_or_str(user_id),
        "strategy_id": payload.strategy_id,
        "instrument_id": payload.instrument_id,
        "timeframe": payload.timeframe,
        "period": f"{payload.start_date.isoformat()} to {payload.end_date.isoformat()}",
        "start_date": payload.start_date,
        "end_date": payload.end_date,
        "initial_capital": _decimal(payload.capital),
        "final_capital": _decimal(service_response.final_capital),
        "net_profit": _decimal(metrics.get("net_profit")),
        "max_drawdown": _decimal(metrics.get("max_drawdown")),
        "sharpe_ratio": _decimal(metrics.get("sharpe_ratio")),
        "win_rate": win_rate,
        "total_trades": total_trades,
        "winning_trades": winning_trades,
        "losing_trades": losing_trades,
        "profit_factor": _decimal(metrics.get("profit_factor")),
        "return_pct": _decimal(metrics.get("return_pct")),
        "avg_win": _decimal(metrics.get("avg_win")),
        "avg_loss": _decimal(metrics.get("avg_loss")),
        "expectancy": _decimal(metrics.get("expectancy")),
        # asyncpg JSONB raw text() inserts require a JSON-encoded string, not a Python dict.
        # Without this, advanced filter runs fail with: dict object has no attribute encode.
        "advanced_filters": json.dumps(filter_meta.get("advanced_filters")) if filter_meta.get("advanced_filters") is not None else None,
        "filter_summary": filter_meta.get("filter_summary"),
        "candles_before_filter": filter_meta.get("candles_before_filter"),
        "candles_after_filter": filter_meta.get("candles_after_filter"),
        "filter_reduction_pct": _decimal(filter_meta.get("filter_reduction_pct"), Decimal("0")) if filter_meta.get("filter_reduction_pct") is not None else None,
        "account_currency": getattr(service_response.result, "account_currency", None),
        "currency_symbol": getattr(service_response.result, "currency_symbol", None),
        "quantity_mode": getattr(service_response.result, "quantity_mode", None),
        "runtime_config_snapshot": json.dumps(getattr(service_response, "runtime_config", None)) if getattr(service_response, "runtime_config", None) is not None else None,
        "instrument_spec_snapshot": json.dumps(getattr(service_response, "instrument_spec", None)) if getattr(service_response, "instrument_spec", None) is not None else None,
        "professional_summary": json.dumps(getattr(service_response.result, "summary", {}) or {}),
        "risk_engine_version": (getattr(service_response.result, "summary", {}) or {}).get("risk_engine_version"),
        "pnl_engine_version": (getattr(service_response.result, "summary", {}) or {}).get("pnl_engine_version"),
        "warnings": json.dumps(service_response.warnings or []),
        "rejected_trade_count": int(getattr(service_response, "rejected_trade_count", 0) or 0),
        "rejection_reasons": json.dumps(getattr(service_response, "rejection_reasons", {}) or {}),
        "trade_details": json.dumps(_trade_json_from_service(service_response)),
        "status": "completed",
    }

    column_meta = await _table_columns_meta(db, "performance_metrics")
    available_columns = [meta["column_name"] for meta in column_meta]
    missing_columns = sorted(set(insert_values.keys()) - set(available_columns))
    if missing_columns:
        logger.warning(
            "performance_metrics is missing columns %s; persisting compatible subset only",
            ", ".join(missing_columns),
        )

    if not available_columns:
        raise HTTPException(
            status_code=500,
            detail={
                "code": "DB_MIGRATION_REQUIRED",
                "message": "performance_metrics table is missing required backtest columns. Run scripts/backtest_performance_metrics_safe_migration.sql and retry.",
                "migration": "scripts/backtest_performance_metrics_safe_migration.sql",
            },
        )

    final_insert_values = {column: insert_values[column] for column in available_columns if column in insert_values}
    class _PayloadShim:
        pass
    payload_shim = _PayloadShim()
    payload_shim.user_id = user_id
    payload_shim.strategy_id = payload.strategy_id
    payload_shim.instrument_id = payload.instrument_id
    payload_shim.timeframe = payload.timeframe
    payload_shim.start_date = payload.start_date
    payload_shim.end_date = payload.end_date
    payload_shim.capital = payload.capital

    for meta in column_meta:
        column = meta["column_name"]
        if column in final_insert_values:
            continue
        if str(meta.get("is_nullable", "YES")).upper() == "NO" and not meta.get("column_default"):
            fallback = _default_perf_metric_value(
                column,
                payload_shim,
                service_response,
                metrics,
                winning_trades,
                losing_trades,
                total_trades,
            )
            if fallback is None:
                data_type = str(meta.get("data_type") or meta.get("udt_name") or "").lower()
                if any(token in data_type for token in ["int", "numeric", "double", "real", "decimal"]):
                    fallback = 0
                elif data_type == "date":
                    fallback = payload.start_date
                elif "time" in data_type:
                    fallback = datetime.utcnow()
                elif meta.get("udt_name") == "uuid":
                    fallback = None
                else:
                    fallback = ""
            if fallback is not None:
                final_insert_values[column] = fallback

    columns_sql = ", ".join(final_insert_values.keys())
    values_sql = ", ".join(f":{column}" for column in final_insert_values.keys())
    await db.execute(
        text(f"INSERT INTO performance_metrics ({columns_sql}) VALUES ({values_sql})"),
        final_insert_values,
    )
    await db.flush()

    trades_fk_value = await _resolve_backtest_fk_value(db, "trades", backtest_id)
    equity_fk_value = await _resolve_backtest_fk_value(db, "equity_curve", backtest_id)
    pnl_fk_value = await _resolve_backtest_fk_value(db, "pnl_calendar", backtest_id)

    if await _table_exists(db, "metrics"):
        try:
            async with db.begin_nested():
                for metric_name in ["net_profit", "win_rate", "max_drawdown", "sharpe_ratio", "profit_factor"]:
                    await db.execute(
                        text(
                            """
                            INSERT INTO metrics (id, backtest_id, name, value)
                            VALUES (:id, :backtest_id, :name, :value)
                            """
                        ),
                        {
                            "id": str(uuid4()),
                            "backtest_id": backtest_id,
                            "name": metric_name,
                            "value": _to_float(metrics.get(metric_name)),
                        },
                    )
        except Exception as exc:
            logger.warning("Failed to persist metrics rows for backtest %s: %s", backtest_id, exc)

    if service_response.result.trades:
        try:
            async with db.begin_nested():
                trade_columns_meta = await _table_columns_meta(db, "trades")
                trade_columns = {meta["column_name"] for meta in trade_columns_meta}
                for trade in service_response.result.trades:
                    base_values = {
                        "id": _uuid_or_bigint_for_column(next((meta for meta in trade_columns_meta if meta["column_name"] == "id"), {"data_type": "bigint", "udt_name": "int8"})),
                        "backtest_id": trades_fk_value,
                        "instrument_id": payload.instrument_id,
                        "entry_time": _ensure_aware_datetime(trade.entry_datetime),
                        "exit_time": _ensure_aware_datetime(trade.exit_datetime),
                        "side": trade.direction,
                        "quantity": _decimal(getattr(trade, "quantity", 0.0)),
                        "lot_size": _decimal(getattr(trade, "lot_size", None)),
                        "account_currency": getattr(trade, "account_currency", None),
                        "currency_symbol": getattr(trade, "currency_symbol", None),
                        "asset_class": getattr(trade, "asset_class", None),
                        "quantity_mode": getattr(trade, "quantity_mode", None),
                        "entry_price": _decimal(trade.entry_price),
                        "exit_price": _decimal(trade.exit_price),
                        "pnl": _decimal(trade.pnl),
                        "exit_type": trade.exit_reason,
                        "exit_reason": trade.exit_reason,
                        "actual_risk_amount": _decimal(getattr(trade, "actual_risk_amount", None)),
                        "risk_ticks": _decimal(getattr(trade, "risk_ticks", None)),
                        "risk_pips": _decimal(getattr(trade, "risk_pips", None)),
                        "reward_ticks": _decimal(getattr(trade, "reward_ticks", None)),
                        "expected_reward_amount": _decimal(getattr(trade, "expected_reward_amount", None)),
                        "sl_mode": getattr(trade, "sl_mode", None),
                        "position_size_mode": getattr(trade, "position_size_mode", None),
                        "runtime_config_snapshot": json.dumps(getattr(trade, "runtime_config_snapshot", None)) if getattr(trade, "runtime_config_snapshot", None) is not None else None,
                        "instrument_spec_snapshot": json.dumps(getattr(trade, "instrument_spec_snapshot", None)) if getattr(trade, "instrument_spec_snapshot", None) is not None else None,
                        "lifecycle_events": json.dumps(getattr(trade, "lifecycle_events", None) or []),
                    }
                    base_values.update({key: (_decimal(value) if isinstance(value, (int, float)) else value) for key, value in _trade_transparency_values(trade).items()})
                    insert_trade_values = _build_trade_insert_values(
                        trade_columns_meta,
                        base_values={key: value for key, value in base_values.items() if key in trade_columns},
                        backtest_id=backtest_id,
                        trades_fk_value=trades_fk_value,
                        user_id=user_id,
                        payload=payload,
                        trade=trade,
                    )
                    if not insert_trade_values:
                        continue
                    columns_sql = ", ".join(insert_trade_values.keys())
                    values_sql = ", ".join(f":{column}" for column in insert_trade_values.keys())
                    await db.execute(text(f"INSERT INTO trades ({columns_sql}) VALUES ({values_sql})"), insert_trade_values)
        except Exception as exc:
            logger.exception("Failed to persist trades for backtest %s. Report will use performance_metrics.trade_details fallback when available. Error: %s", backtest_id, exc)

    if service_response.result.equity_curve:
        try:
            async with db.begin_nested():
                base = datetime.combine(payload.start_date, time(0, 0, 0))
                for idx, equity in enumerate(service_response.result.equity_curve):
                    db.add(
                        EquityCurve(
                            backtest_id=equity_fk_value,
                            timestamp=_ensure_aware_datetime(base + timedelta(minutes=idx)),
                            equity=_decimal(equity),
                        )
                    )
        except Exception as exc:
            logger.warning("Failed to persist equity curve for backtest %s: %s", backtest_id, exc)

    if service_response.result.trades:
        try:
            pnl_map: dict[date, Decimal] = {}
            for trade in service_response.result.trades:
                if not trade.exit_datetime:
                    continue
                day = trade.exit_datetime.date()
                pnl_map[day] = pnl_map.get(day, Decimal("0")) + _decimal(trade.pnl)

            if pnl_map:
                async with db.begin_nested():
                    for day, pnl in pnl_map.items():
                        db.add(PnLCalendar(backtest_id=pnl_fk_value, date=day, pnl=pnl))
        except Exception as exc:
            logger.warning("Failed to persist pnl calendar for backtest %s: %s", backtest_id, exc)

    return backtest_id



def _runtime_section_rows(runtime_config: dict | None) -> list[list[Any]]:
    cfg = runtime_config if isinstance(runtime_config, dict) else {}
    if not cfg:
        return [["Snapshot Status", "Runtime settings snapshot was not available for this older backtest."]]
    risk = cfg.get("risk") if isinstance(cfg.get("risk"), dict) else {}
    sl_tp = cfg.get("sl_tp") if isinstance(cfg.get("sl_tp"), dict) else {}
    execution = cfg.get("execution") if isinstance(cfg.get("execution"), dict) else {}
    tm = cfg.get("trade_management") if isinstance(cfg.get("trade_management"), dict) else {}
    params = cfg.get("strategy_params") if isinstance(cfg.get("strategy_params"), dict) else {}

    def yn(value: Any) -> str:
        return "ON" if bool(value) else "OFF"

    rows: list[list[Any]] = [
        ["Risk · Initial Capital", risk.get("initial_capital")],
        ["Risk · Capital Risk %", risk.get("risk_percent")],
        ["Risk · Position Size Mode", risk.get("position_size_mode")],
        ["Risk · Fixed Lot Size", risk.get("fixed_lot_size")],
        ["Risk · Max Lot Cap", risk.get("max_lot_cap")],
        ["Risk · Max Quantity Cap", risk.get("max_quantity_cap")],
        ["SL/TP · RR Ratio", sl_tp.get("rr_ratio")],
        ["SL/TP · SL Mode", sl_tp.get("sl_mode")],
        ["SL/TP · Fixed Price Risk %", sl_tp.get("fixed_price_risk_pct")],
        ["SL/TP · ATR Period", sl_tp.get("atr_period")],
        ["SL/TP · ATR Multiplier", sl_tp.get("atr_multiplier")],
        ["SL/TP · Swing Lookback", sl_tp.get("swing_lookback")],
        ["Execution · Entry Mode", execution.get("entry_mode")],
        ["Execution · Exit On Opposite Signal", yn(execution.get("exit_on_opposite_signal"))],
        ["Execution · Allow Long", yn(execution.get("allow_long"))],
        ["Execution · Allow Short", yn(execution.get("allow_short"))],
        ["Execution · Max Trades Per Day", execution.get("max_trades_per_day")],
        ["Execution · Max Open Positions", execution.get("max_open_positions")],
        ["Trade Mgmt · Break Even Enabled", yn(tm.get("break_even_enabled"))],
        ["Trade Mgmt · Break Even Trigger R", tm.get("break_even_trigger_r")],
        ["Trade Mgmt · Trailing Stop Enabled", yn(tm.get("trailing_enabled"))],
        ["Trade Mgmt · Trailing Mode", tm.get("trailing_mode")],
        ["Trade Mgmt · Trail Start R", tm.get("trail_start_r")],
        ["Trade Mgmt · Trail ATR Multiplier", tm.get("trail_atr_multiplier")],
        ["Trade Mgmt · Partial Exit Enabled", yn(tm.get("partial_exit_enabled"))],
        ["Trade Mgmt · Partial Exit At R", tm.get("partial_exit_at_r")],
        ["Trade Mgmt · Partial Exit Percent", tm.get("partial_exit_percent")],
    ]
    for key, value in params.items():
        rows.append([f"Strategy Params · {key}", value])
    return rows


def _runtime_summary_text(runtime_config: dict | None, filter_summary: str | None = None) -> str:
    cfg = runtime_config if isinstance(runtime_config, dict) else {}
    if not cfg:
        base = "Runtime snapshot unavailable"
    else:
        risk = cfg.get("risk") if isinstance(cfg.get("risk"), dict) else {}
        sl_tp = cfg.get("sl_tp") if isinstance(cfg.get("sl_tp"), dict) else {}
        tm = cfg.get("trade_management") if isinstance(cfg.get("trade_management"), dict) else {}
        parts: list[str] = []
        risk_value = risk.get("risk_percent")
        if risk_value is not None:
            try:
                rv = float(risk_value)
                risk_label = f"Risk {rv * 100:g}%" if abs(rv) <= 1 else f"Risk {rv:g}%"
            except Exception:
                risk_label = f"Risk {risk_value}"
            parts.append(risk_label)
        rr_value = sl_tp.get("rr_ratio")
        if rr_value is not None:
            parts.append(f"RR 1:{rr_value}")
        sl_mode = sl_tp.get("sl_mode")
        if sl_mode:
            parts.append(f"{str(sl_mode).replace('_', ' ').title()} SL")
        pos_mode = risk.get("position_size_mode")
        if pos_mode:
            if str(pos_mode).upper() == "FIXED_LOT" and risk.get("fixed_lot_size") is not None:
                parts.append(f"Fixed Lot {risk.get('fixed_lot_size')}")
            else:
                parts.append(str(pos_mode).replace('_', ' ').title())
        if tm.get("break_even_enabled"):
            parts.append("Breakeven ON")
        if tm.get("trailing_enabled"):
            parts.append("Trail ON")
        if tm.get("partial_exit_enabled"):
            parts.append("Partial Exit ON")
        base = " · ".join(parts) if parts else "Runtime snapshot captured"
    if filter_summary:
        base = f"{base} · Advanced Filters: {filter_summary}"
    return base

def _build_detail_export_frames(detail: dict):
    summary = detail.get("summary", {}) if isinstance(detail, dict) else {}
    metrics_rows = [
        ["Backtest ID", summary.get("id")],
        ["Strategy", summary.get("strategy_name")],
        ["Instrument", summary.get("instrument_symbol")],
        ["Asset Class", summary.get("asset_class")],
        ["Account Currency", summary.get("account_currency")],
        ["Quantity Mode", summary.get("quantity_mode")],
        ["Position Size Mode", summary.get("position_size_mode")],
        ["SL Mode", summary.get("sl_mode")],
        ["RR Ratio", summary.get("rr_ratio")],
        ["Risk %", summary.get("risk_percent")],
        ["Timeframe", summary.get("timeframe")],
        ["Initial Capital", summary.get("initial_capital")],
        ["Final Capital", summary.get("final_capital")],
        ["Net Profit", summary.get("net_profit")],
        ["Return %", summary.get("return_pct")],
        ["Win Rate", summary.get("win_rate")],
        ["Sharpe Ratio", summary.get("sharpe_ratio")],
        ["Max Drawdown", summary.get("max_drawdown")],
        ["Profit Factor", summary.get("profit_factor")],
        ["Avg Win", summary.get("avg_win")],
        ["Avg Loss", summary.get("avg_loss")],
        ["Expectancy", summary.get("expectancy")],
        ["Total Trades", summary.get("total_trades")],
        ["Winning Trades", summary.get("winning_trades")],
        ["Losing Trades", summary.get("losing_trades")],
        ["Advanced Filters", summary.get("filter_summary") or "Not used"],
        ["Candles Before Filter", summary.get("candles_before_filter")],
        ["Candles After Filter", summary.get("candles_after_filter")],
        ["Filter Reduction %", summary.get("filter_reduction_pct")],
        ["Created At", summary.get("created_at")],
    ]
    runtime_config = _jsonish(summary.get("runtime_config_snapshot"))
    if not isinstance(runtime_config, dict):
        runtime_config = {}
    runtime_rows = _runtime_section_rows(runtime_config)
    runtime_rows.append(["Advanced Filters", summary.get("filter_summary") or "Not used"])
    metrics_df = pd.DataFrame(metrics_rows, columns=["Metric", "Value"])
    runtime_df = pd.DataFrame(runtime_rows, columns=["Runtime Setting", "Value"])
    trades_df = pd.DataFrame(detail.get("trades", []))
    equity_df = pd.DataFrame(detail.get("equity_curve", []))
    pnl_df = pd.DataFrame(detail.get("pnl_calendar", []))
    return metrics_df, runtime_df, trades_df, equity_df, pnl_df


async def _detail_payload_for_export(backtest_id: str, db: AsyncSession, current_user: dict) -> dict:
    response = await get_backtest_detail(backtest_id=backtest_id, db=db, current_user=current_user)
    payload = response.get("data") if isinstance(response, dict) and "data" in response else response
    return payload


def _autosize_excel_sheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="2F1B57")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(v) for v in values) + 2, 42) if values else 12
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)


def _build_trade_analysis_frames(detail: dict):
    trades = detail.get("trades", []) if isinstance(detail, dict) else []
    if not trades:
        empty = pd.DataFrame(columns=["label", "value"])
        return empty, empty, empty

    trades_df = pd.DataFrame(trades)
    if not trades_df.empty:
        trades_df["entry_time"] = pd.to_datetime(trades_df.get("entry_time"), errors="coerce")
        trades_df["exit_time"] = pd.to_datetime(trades_df.get("exit_time"), errors="coerce")
        trades_df["pnl"] = pd.to_numeric(trades_df.get("pnl"), errors="coerce").fillna(0.0)
        trades_df["quantity"] = pd.to_numeric(trades_df.get("quantity"), errors="coerce").fillna(0)
        trades_df["duration_minutes"] = ((trades_df["exit_time"] - trades_df["entry_time"]).dt.total_seconds() / 60.0).fillna(0.0)

    side_breakdown = (
        trades_df.groupby("side", dropna=False)
        .agg(trades=("id", "count"), total_pnl=("pnl", "sum"), avg_pnl=("pnl", "mean"))
        .reset_index()
    ) if not trades_df.empty else pd.DataFrame(columns=["side", "trades", "total_pnl", "avg_pnl"])

    daily_trades = pd.DataFrame(columns=["date", "trade_count", "daily_pnl"])
    if not trades_df.empty and "exit_time" in trades_df.columns:
        temp = trades_df.copy()
        temp["date"] = temp["exit_time"].dt.date
        daily_trades = temp.groupby("date").agg(trade_count=("id", "count"), daily_pnl=("pnl", "sum")).reset_index()

    highlights = pd.DataFrame([
        ["Largest Win", float(trades_df["pnl"].max()) if not trades_df.empty else 0.0],
        ["Largest Loss", float(trades_df["pnl"].min()) if not trades_df.empty else 0.0],
        ["Average Duration (min)", float(trades_df["duration_minutes"].mean()) if not trades_df.empty else 0.0],
        ["Median Duration (min)", float(trades_df["duration_minutes"].median()) if not trades_df.empty else 0.0],
        ["Average Position Size", float(trades_df["quantity"].mean()) if not trades_df.empty else 0.0],
    ], columns=["Metric", "Value"])

    return side_breakdown, daily_trades, highlights


@router.get("/{backtest_id}/export/excel")
async def export_backtest_excel(backtest_id: str, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    detail = await _detail_payload_for_export(backtest_id, db, current_user)
    metrics_df, runtime_df, trades_df, equity_df, pnl_df = _build_detail_export_frames(detail)
    side_df, daily_trades_df, highlights_df = _build_trade_analysis_frames(detail)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        metrics_df.to_excel(writer, sheet_name="Summary", index=False)
        highlights_df.to_excel(writer, sheet_name="Highlights", index=False)
        runtime_df.to_excel(writer, sheet_name="Runtime Settings", index=False)
        trades_df.to_excel(writer, sheet_name="Trades", index=False)
        side_df.to_excel(writer, sheet_name="Trade Breakdown", index=False)
        daily_trades_df.to_excel(writer, sheet_name="Daily Trades", index=False)
        equity_df.to_excel(writer, sheet_name="Equity Curve", index=False)
        pnl_df.to_excel(writer, sheet_name="PnL Calendar", index=False)
    output.seek(0)
    workbook = load_workbook(output)
    for sheet in workbook.worksheets:
        _autosize_excel_sheet(sheet)
        sheet.freeze_panes = "A2"
    formatted = BytesIO()
    workbook.save(formatted)
    formatted.seek(0)
    filename = f"backtest-{backtest_id}-full-report.xlsx"
    return StreamingResponse(formatted, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/{backtest_id}/export/pdf")
async def export_backtest_pdf(backtest_id: str, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    request_id = str(uuid4())
    try:
        from ...services.reports.backtest_pdf_report import build_backtest_pdf

        detail = await _detail_payload_for_export(backtest_id, db, current_user)
        output, filename = build_backtest_pdf(detail)
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("PDF export failed for backtest %s request_id=%s", backtest_id, request_id)
        raise HTTPException(
            status_code=500,
            detail=f"PDF export failed. Please retry or contact support with request ID {request_id}.",
        ) from exc


@router.get("/config")
async def get_backtest_config(
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    entitlements: dict = Depends(get_user_entitlements),
):
    user_id = current_user["user_id"]
    uid = as_uuid_or_str(user_id)

    strategies = (
        await db.execute(
            select(Strategy)
            .where(or_(Strategy.visibility == "PUBLIC", Strategy.created_by == uid))
            .order_by(Strategy.created_at.desc())
        )
    ).scalars().all()
    instruments = (await db.execute(select(Instrument).order_by(Instrument.symbol.asc()))).scalars().all()
    timeframes = (
        await db.execute(select(MarketData.timeframe).distinct().order_by(MarketData.timeframe.asc()))
    ).scalars().all()

    try:

        capacity = await CreditManagementService.get_credit_capacity(db, str(user_id), for_update=False)
    except Exception:
        capacity = {
            "wallet_balance": 0,
            "included_balance": 0,
            "total_available": 0,
            "subscription_state": "NONE",
            "subscription_id": None,
            "refill_applied": False,
            "next_refill_at": None,
        }

    return success_response(
        {
            "strategies": [{"id": str(s.id), "name": s.name} for s in strategies],
            "instruments": [
                {
                    "id": i.id,
                    "symbol": i.symbol,
                    "exchange": i.exchange,
                    "market": i.market,
                    "instrument_type": i.instrument_type,
                }
                for i in instruments
            ],
            "timeframes": [tf for tf in timeframes if tf],
            "credits": {
                "balance": _to_float(capacity.get("total_available") or 0),
                "current_balance": _to_float(capacity.get("total_available") or 0),
                "wallet_balance": _to_int(capacity.get("wallet_balance") or 0),
                "included": _to_int(capacity.get("included_balance") or 0),
                "included_balance": _to_int(capacity.get("included_balance") or 0),
                "total_available": _to_int(capacity.get("total_available") or 0),
                "subscription_state": capacity.get("subscription_state"),
                "next_refill_at": capacity.get("next_refill_at").isoformat() if capacity.get("next_refill_at") else None,
                "deduction_order": ["subscription", "wallet"],
            },
            "limits": {
                "max_backtests_per_day": _to_int(entitlements.get("features", {}).get("backtests_per_day", 5)),
                "max_date_range_days": _to_int(entitlements.get("features", {}).get("max_date_range_days", 30)),
            },
        }
    )


@router.get("/timeframes")
async def get_backtest_timeframes(
    instrument_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(MarketData.timeframe).distinct().order_by(MarketData.timeframe.asc())
    if instrument_id is not None:
        stmt = stmt.where(MarketData.instrument_id == instrument_id)
    rows = (await db.execute(stmt)).scalars().all()
    return success_response({"timeframes": [tf for tf in rows if tf]})


@router.get("/data-availability")
async def get_data_availability(
    instrument_id: int,
    timeframe: str,
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    parsed_start_date = _parse_date_param(start_date, "start_date")
    parsed_end_date = _parse_date_param(end_date, "end_date")

    summary = (
        await db.execute(
            select(
                func.min(MarketData.timestamp).label("min_ts"),
                func.max(MarketData.timestamp).label("max_ts"),
                func.count().label("count"),
            ).where(MarketData.instrument_id == instrument_id, MarketData.timeframe == timeframe)
        )
    ).first()

    if not summary or summary.min_ts is None:
        return success_response(
            {
                "instrument_id": instrument_id,
                "timeframe": timeframe,
                "status": "error",
                "available": False,
                "coverage_status": "BLOCKED",
                "message": "Market data is missing for this instrument/timeframe/date range. Ask admin to import missing candles.",
                "candle_count": 0,
                "total_candles": 0,
                "matched_candles": 0,
                "min_timestamp": None,
                "max_timestamp": None,
                "requested_candle_count": 0,
            }
        )

    requested_count = None
    range_min = None
    range_max = None
    missing_before = False
    missing_after = False
    coverage_status = "AVAILABLE"
    message = "Market data available for selected range."

    if parsed_start_date and parsed_end_date:
        guard = await _get_market_data_availability_guard(db, instrument_id, timeframe, parsed_start_date, parsed_end_date)
        requested_count = guard["record_count"]
        range_min = guard.get("available_start")
        range_max = guard.get("available_end")
        missing_before = bool(guard.get("missing_before"))
        missing_after = bool(guard.get("missing_after"))
        coverage_status = "BLOCKED" if guard.get("blocked") else "AVAILABLE"
        message = str(guard.get("message") or message)

    return success_response(
        {
            "instrument_id": instrument_id,
            "timeframe": timeframe,
            "status": "ok" if coverage_status == "AVAILABLE" else "error",
            "available": coverage_status == "AVAILABLE",
            "coverage_status": coverage_status,
            "message": message,
            "candle_count": _to_int(summary.count),
            "total_candles": _to_int(summary.count),
            "matched_candles": _to_int(requested_count, 0),
            "min_timestamp": summary.min_ts.isoformat() if summary.min_ts else None,
            "max_timestamp": summary.max_ts.isoformat() if summary.max_ts else None,
            "available_start": range_min.isoformat() if range_min else None,
            "available_end": range_max.isoformat() if range_max else None,
            "requested_start": datetime.combine(parsed_start_date, time.min).isoformat() if parsed_start_date else None,
            "requested_end": datetime.combine(parsed_end_date, time.max).isoformat() if parsed_end_date else None,
            "missing_before": missing_before,
            "missing_after": missing_after,
            "requested_candle_count": _to_int(requested_count, 0),
        }
    )


@router.post("/cost-preview")
async def preview_backtest_cost(
    payload: BacktestCostPreviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    entitlements: dict = Depends(get_user_entitlements),
):
    plan_code = str(entitlements.get("plan_code") or "").upper() if entitlements else None

    advanced_filter_preview = await _advanced_filter_preview(
        db,
        instrument_id=payload.instrument_id,
        timeframe=payload.timeframe,
        start_date=payload.start_date,
        end_date=payload.end_date,
        advanced_filters=payload.advanced_filters,
    )

    filtered_candle_override = None
    candle_count_mode_override = None
    if advanced_filter_preview is not None:
        filtered_candle_override = _to_int(advanced_filter_preview.get("total_candles_after_filter"), 0)
        candle_count_mode_override = "filtered_actual"

    estimate = await _quote_backtest_cost(
        db,
        plan_code=plan_code,
        strategy_id=payload.strategy_id,
        instrument_id=payload.instrument_id,
        timeframe=payload.timeframe,
        start_date=payload.start_date,
        end_date=payload.end_date,
        use_actual_candle_count=bool(payload.instrument_id),
        candle_count_override=filtered_candle_override,
        candle_count_mode_override=candle_count_mode_override,
        advanced_filters=payload.advanced_filters,
    )

    capacity = await CreditManagementService.get_credit_capacity(db, str(current_user["user_id"]), for_update=False)
    total_available = int(capacity.get("total_available") or 0)
    can_run = float(total_available) >= estimate["total_cost"]

    warnings: list[str] = []
    if advanced_filter_preview is not None:
        warnings.extend([str(item) for item in advanced_filter_preview.get("warnings") or [] if str(item).strip()])

    before_count = _to_int(
        advanced_filter_preview.get("total_candles_before_filter") if advanced_filter_preview else estimate.get("breakdown", {}).get("candle_count"),
        0,
    )
    after_count = _to_int(
        advanced_filter_preview.get("total_candles_after_filter") if advanced_filter_preview else before_count,
        before_count,
    )
    removed = _to_int(
        advanced_filter_preview.get("candles_removed") if advanced_filter_preview else 0,
        0,
    )
    reduction_pct = float(
        advanced_filter_preview.get("filter_reduction_pct") if advanced_filter_preview else 0.0
        or 0.0
    )

    data_coverage = {
        "status": (advanced_filter_preview or {}).get("status", "ok"),
        "range_start": payload.start_date.isoformat(),
        "range_end": payload.end_date.isoformat(),
        "total_candles": before_count,
        "filtered_candles": after_count,
        "candles_removed": removed,
        "filter_reduction_pct": reduction_pct,
    }

    response_payload = {
        "instrument_id": payload.instrument_id,
        "timeframe": payload.timeframe,
        "date_range": {
            "start_date": payload.start_date.isoformat(),
            "end_date": payload.end_date.isoformat(),
        },
        "total_cost": estimate["total_cost"],
        "estimated_run_cost": estimate["total_cost"],
        "estimated_candles": estimate.get("estimated_candles") or after_count or before_count,
        "credit_cost": estimate["total_cost"],
        "credit_balance": float(total_available),
        "has_enough_credits": can_run,
        "pricing_rule": estimate.get("pricing_rule") or estimate.get("breakdown", {}).get("rule_set_name"),
        "breakdown": estimate["breakdown"],
        "pricing_rule_set": {
            "id": estimate.get("breakdown", {}).get("rule_set_id"),
            "name": estimate.get("breakdown", {}).get("rule_set_name"),
            "version": estimate.get("breakdown", {}).get("pricing_version"),
        },
        "current_balance": float(total_available),
        "wallet_balance": int(capacity.get("wallet_balance") or 0),
        "included_balance": int(capacity.get("included_balance") or 0),
        "balances": {
            "wallet_balance": int(capacity.get("wallet_balance") or 0),
            "included_balance": int(capacity.get("included_balance") or 0),
            "total_available": int(capacity.get("total_available") or 0),
        },
        "subscription_state": capacity.get("subscription_state"),
        "can_run": can_run and data_coverage["status"] != "error",
        "cost_feasible": can_run,
        "data_coverage": data_coverage,
        "advanced_filters": advanced_filter_preview or {
            "enabled": False,
            "summary": "Advanced filters disabled",
            "total_candles_before_filter": before_count,
            "total_candles_after_filter": before_count,
            "candles_removed": 0,
            "filter_reduction_pct": 0.0,
            "warnings": [],
        },
        "warnings": warnings,
    }

    return success_response(response_payload)


@router.post("/run", status_code=status.HTTP_202_ACCEPTED)
async def run_backtest(
    payload: BacktestRunRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    entitlements: dict = Depends(get_user_entitlements),
):
    strategy = await db.get(Strategy, payload.strategy_id)
    if not strategy:
        raise HTTPException(status_code=404, detail="Strategy not found")

    instrument = await db.get(Instrument, payload.instrument_id)
    if not instrument:
        raise HTTPException(status_code=404, detail="Instrument not found")

    availability = await _get_market_data_availability_guard(
        db,
        instrument_id=payload.instrument_id,
        timeframe=payload.timeframe,
        start_date=payload.start_date,
        end_date=payload.end_date,
    )
    if availability.get("blocked"):
        _raise_market_data_unavailable(availability)

    plan_code = str(entitlements.get("plan_code") or "").upper() if entitlements else None
    advanced_filter_preview = await _advanced_filter_preview(
        db,
        instrument_id=payload.instrument_id,
        timeframe=payload.timeframe,
        start_date=payload.start_date,
        end_date=payload.end_date,
        advanced_filters=payload.advanced_filters,
    )
    filtered_candle_override = None
    candle_count_mode_override = None
    if advanced_filter_preview is not None:
        filtered_candle_override = _to_int(advanced_filter_preview.get("total_candles_after_filter"), 0)
        candle_count_mode_override = "filtered_actual"

    estimate = await _quote_backtest_cost(
        db,
        plan_code=plan_code,
        strategy_id=payload.strategy_id,
        instrument_id=payload.instrument_id,
        timeframe=payload.timeframe,
        start_date=payload.start_date,
        end_date=payload.end_date,
        use_actual_candle_count=True,
        candle_count_override=filtered_candle_override,
        candle_count_mode_override=candle_count_mode_override,
        advanced_filters=payload.advanced_filters,
    )
    cost = Decimal(str(estimate["total_cost"]))


    capacity = await CreditManagementService.get_credit_capacity(db, str(current_user["user_id"]), for_update=False)
    total_available = int(capacity.get("total_available") or 0)
    if total_available < _to_int(cost):
        raise HTTPException(
            status_code=402,
            detail={
                "code": "INSUFFICIENT_CREDITS",
                "message": "Insufficient credits for backtest. Upgrade plan or top-up wallet credits to continue.",
                "needed": _to_int(cost),
                "balance": float(total_available),
                "wallet_balance": int(capacity.get("wallet_balance") or 0),
                "included_balance": int(capacity.get("included_balance") or 0),
                "subscription_state": capacity.get("subscription_state"),
                "action": {
                    "upgrade_url": "/pricing",
                    "topup_url": "/credits",
                },
            },
        )
    if not await _column_exists(db, "job_status", "debit_txn_id"):
        raise HTTPException(
            status_code=500,
            detail={
                "code": "DB_MIGRATION_REQUIRED",
                "message": "Database schema is missing job_status.debit_txn_id required for backtest credit tracking. Run scripts/backtest_job_status_credit_link_safe_migration.sql and retry.",
                "migration": "scripts/backtest_job_status_credit_link_safe_migration.sql",
            },
        )

    if not await _column_exists(db, "performance_metrics", "instrument_id"):
        logger.warning(
            "performance_metrics.instrument_id missing; history/detail pages will be limited until scripts/backtest_performance_metrics_safe_migration.sql is applied"
        )

    job_id = str(uuid4())
    job = JobStatus(
        id=job_id,
        user_id=as_uuid_or_str(current_user["user_id"]),
        job_type="backtest",
        status="running",
        progress=10,
        message="Running backtest",
        job_data=json.dumps(payload.model_dump(mode="json")),
        started_at=datetime.utcnow(),
    )
    db.add(job)
    await db.flush()

    debit_txn = None
    included_txn = None
    consumption = None
    try:
        consumption = await CreditManagementService.consume_credits_for_backtest(
            db=db,
            user_id=str(current_user["user_id"]),
            total_cost=cost,
            description=(
                f"Backtest run: {getattr(instrument, 'symbol', payload.instrument_id)} {payload.timeframe} "
                f"{payload.start_date.isoformat()} to {payload.end_date.isoformat()} | "
                f"candles: {estimate.get('estimated_candles') or estimate.get('breakdown', {}).get('candle_count') or 0} | "
                f"rule: {estimate.get('pricing_rule') or estimate.get('breakdown', {}).get('rule_set_name') or 'Credit expense rule'}"
            ),
            job_id=job_id,
            auto_commit=False,
        )

        included_txn = consumption.get("included_transaction")
        debit_txn = consumption.get("wallet_transaction")

        if debit_txn is not None:
            job.debit_txn_id = str(debit_txn.id)
        elif included_txn is not None:
            job.debit_txn_id = str(included_txn.id)
        job.progress = 40
        job.message = "Executing strategy"

        service_response = await BacktestService.run_backtest(
            db=db,
            strategy_id=payload.strategy_id,
            instrument_id=payload.instrument_id,
            timeframe=payload.timeframe,
            start_date=payload.start_date,
            end_date=payload.end_date,
            initial_capital=payload.capital,
            advanced_filters=payload.advanced_filters,
            runtime_config=payload.runtime_config,
            strategy_preset_id=payload.strategy_preset_id,
            timeframe_id=payload.timeframe_id,
        )

        trade_df = _trade_df_from_service(service_response)
        equity_df = _equity_df_from_service(service_response, payload.start_date)
        metrics = MetricsCalculator.calculate_all_metrics(
            equity_curve=equity_df,
            trades=trade_df,
            initial_capital=_to_float(payload.capital),
        )

        backtest_id = await _save_backtest_payload(
            db,
            user_id=str(current_user["user_id"]),
            payload=payload,
            service_response=service_response,
            metrics=metrics,
        )

        for txn in [included_txn, debit_txn]:
            if txn is None:
                continue
            await db.execute(
                update(CreditTransaction)
                .where(CreditTransaction.id == txn.id)
                .values(backtest_id=backtest_id)
            )

        result_data = {
            "backtest_id": backtest_id,
            "strategy_name": service_response.strategy_name,
            "instrument_symbol": service_response.instrument_symbol,
            "timeframe": service_response.timeframe,
            "start_date": service_response.start_date.isoformat(),
            "end_date": service_response.end_date.isoformat(),
            "initial_capital": _to_float(service_response.initial_capital),
            "final_capital": _to_float(service_response.final_capital),
            "net_profit": _to_float(metrics.get("net_profit")),
            "max_drawdown": _to_float(metrics.get("max_drawdown")),
            "sharpe_ratio": _to_float(metrics.get("sharpe_ratio")),
            "win_rate": _to_float(metrics.get("win_rate")),
            "profit_factor": _to_float(metrics.get("profit_factor")),
            "return_pct": _to_float(metrics.get("return_pct")),
            "avg_win": _to_float(metrics.get("avg_win")),
            "avg_loss": _to_float(metrics.get("avg_loss")),
            "expectancy": _to_float(metrics.get("expectancy")),
            "total_trades": _to_int(service_response.total_trades),
            "credit_cost": _to_float(cost),
            "included_credits_used": int((consumption or {}).get("effective_included_debited") or 0),
            "wallet_credits_used": int((consumption or {}).get("effective_wallet_debited") or 0),
            "charge_idempotent": bool((consumption or {}).get("idempotent", False)),
            "included_debit_transaction_id": str(included_txn.id) if included_txn is not None else None,
            "debit_transaction_id": str(debit_txn.id) if debit_txn is not None else None,
            "pricing": estimate.get("breakdown", {}),
            "advanced_filters": _filter_meta_from_impact(service_response.advanced_filter_impact, payload.advanced_filters),
            "advanced_filter_impact": service_response.advanced_filter_impact,
            "warnings": service_response.warnings or [],
            "rejected_trade_count": int(getattr(service_response, "rejected_trade_count", 0) or 0),
            "rejection_reasons": getattr(service_response, "rejection_reasons", {}) or {},
            "risk_engine_version": (getattr(service_response.result, "summary", {}) or {}).get("risk_engine_version"),
            "pnl_engine_version": (getattr(service_response.result, "summary", {}) or {}).get("pnl_engine_version"),
            "account_currency": getattr(service_response.result, "account_currency", None),
            "currency_symbol": getattr(service_response.result, "currency_symbol", None),
            "asset_class": (getattr(service_response.result, "summary", {}) or {}).get("asset_class"),
            "quantity_mode": getattr(service_response.result, "quantity_mode", None),
            "position_size_mode": (getattr(service_response.result, "summary", {}) or {}).get("position_size_mode"),
            "sl_mode": (getattr(service_response.result, "summary", {}) or {}).get("sl_mode"),
            "rr_ratio": (getattr(service_response.result, "summary", {}) or {}).get("rr_ratio"),
            "risk_percent": (getattr(service_response.result, "summary", {}) or {}).get("risk_percent"),
            "avg_lot_size": (getattr(service_response.result, "summary", {}) or {}).get("avg_lot_size"),
            "avg_quantity": (getattr(service_response.result, "summary", {}) or {}).get("avg_quantity"),
            "professional_summary": getattr(service_response.result, "summary", {}) or {},
            "saved": True,
        }

        job.status = "completed"
        job.progress = 100
        job.message = "Backtest completed"
        job.result_data = json.dumps(result_data)
        job.completed_at = datetime.utcnow()
        await db.commit()

        try:
            await NotificationService.create_notification(
                db,
                user_id=str(current_user["user_id"]),
                title="Backtest completed",
                message=f"{service_response.strategy_name} backtest completed on {service_response.instrument_symbol} ({service_response.timeframe}).",
                notification_type="BACKTEST_COMPLETED",
                severity="success",
                entity_type="backtest",
                entity_id=str(backtest_id),
                action_url=f"/backtest-history?backtestId={backtest_id}",
                metadata={"job_id": str(job_id), "net_profit": _to_float(metrics.get("net_profit")), "total_trades": _to_int(service_response.total_trades)},
                auto_commit=True,
            )
        except Exception:
            await db.rollback()
            logger.exception("Failed to create backtest completion notification for %s", backtest_id)

        return success_response(
            {
                "job_id": job_id,
                "status": "completed",
                "backtest_id": backtest_id,
                "result": result_data,
                "credits": {
                    "debited": _to_int(cost),
                    "included_debited": int((consumption or {}).get("effective_included_debited") or 0),
                    "wallet_debited": int((consumption or {}).get("effective_wallet_debited") or 0),
                    "charge_idempotent": bool((consumption or {}).get("idempotent", False)),
                    "included_debit_transaction_id": str(included_txn.id) if included_txn is not None else None,
                    "wallet_debit_transaction_id": str(debit_txn.id) if debit_txn is not None else None,
                    "balance_after": float((consumption or {}).get("wallet_balance_after") or 0),
                    "included_balance_after": int((consumption or {}).get("included_balance_after") or 0),
                    "total_balance_after": float(
                        int((consumption or {}).get("wallet_balance_after") or 0)
                        + int((consumption or {}).get("included_balance_after") or 0)
                    ),
                    "subscription_state": (consumption or {}).get("subscription_state"),
                    "deduction_order": ["subscription", "wallet"],
                },
            },
            "Backtest completed successfully",
        )
    except BacktestError as exc:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(exc))
    except ValueError as exc:
        await db.rollback()
        raise HTTPException(
            status_code=402,
            detail={
                "code": "INSUFFICIENT_CREDITS",
                "message": str(exc) or "Insufficient credits for backtest. Upgrade plan or top-up wallet credits.",
                "needed": _to_int(cost),
                "balance": float((await CreditManagementService.get_credit_capacity(db, str(current_user["user_id"]), for_update=False)).get("total_available") or 0),
                "action": {
                    "upgrade_url": "/pricing",
                    "topup_url": "/credits",
                },
            },
        )
    except Exception as exc:
        logger.exception("Backtest run failed for user %s", current_user["user_id"])
        rollback_ok = True
        try:
            await db.rollback()
        except Exception:
            rollback_ok = False
            logger.exception("Rollback failed after backtest error for job %s", job_id)

        if consumption is not None and not rollback_ok:
            try:
                refund_result = await CreditManagementService.restore_consumed_credits(
                    db=db,
                    user_id=str(current_user["user_id"]),
                    included_amount=0,
                    wallet_amount=0,
                    description=f"Refund for failed backtest job {job_id}",
                    job_id=job_id,
                    auto_commit=False,
                )
                logger.info(
                    "Refunded failed backtest credits for job %s (included=%s, wallet=%s)",
                    job_id,
                    int((refund_result or {}).get("included_refunded") or 0),
                    int((refund_result or {}).get("wallet_refunded") or 0),
                )
            except Exception as refund_exc:
                logger.error("Failed to refund credits for job %s: %s", job_id, refund_exc)

        try:
            await db.execute(
                update(JobStatus)
                .where(JobStatus.id == job_id)
                .values(
                    status="failed",
                    progress=0,
                    message=str(exc),
                    completed_at=datetime.utcnow(),
                )
            )
            await db.commit()
            try:
                await NotificationService.create_notification(
                    db,
                    user_id=str(current_user["user_id"]),
                    title="Backtest failed",
                    message="Your backtest could not be completed. Please review the error and try again.",
                    notification_type="BACKTEST_FAILED",
                    severity="error",
                    entity_type="backtest_job",
                    entity_id=str(job_id),
                    action_url="/backtest",
                    metadata={"job_id": str(job_id), "error": str(exc)},
                    auto_commit=True,
                )
            except Exception:
                await db.rollback()
                logger.exception("Failed to create backtest failure notification for job %s", job_id)
        except Exception:
            await db.rollback()
            logger.exception("Failed to persist failed job status for job %s", job_id)

        detail = str(exc)
        if "performance_metrics" in detail:
            detail = (
                "Backtest execution failed because performance_metrics schema is outdated or has legacy NOT NULL columns. "
                "Run scripts/backtest_performance_metrics_safe_migration.sql and retry."
            )
        raise HTTPException(status_code=500, detail=detail)


@router.get("/")
async def get_backtests(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    return await get_backtest_history(page=page, page_size=page_size, db=db, current_user=current_user)


@router.get("/history")
async def get_backtest_history(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    strategy_id: str | None = Query(default=None),
    instrument_id: int | None = Query(default=None),
    timeframe: str | None = Query(default=None),
    status_filter: str | None = Query(default=None, alias="status"),
    start_date_from: date | None = Query(default=None),
    start_date_to: date | None = Query(default=None),
    min_profit: float | None = Query(default=None),
    max_drawdown: float | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    uid = str(current_user["user_id"])
    filters = [cast(PerformanceMetric.user_id, String) == uid]
    if strategy_id:
        filters.append(PerformanceMetric.strategy_id == strategy_id)
    if instrument_id is not None:
        filters.append(PerformanceMetric.instrument_id == instrument_id)
    if timeframe:
        filters.append(PerformanceMetric.timeframe == timeframe)
    if status_filter:
        filters.append(func.lower(PerformanceMetric.status) == status_filter.lower())
    if start_date_from:
        filters.append(PerformanceMetric.start_date >= start_date_from)
    if start_date_to:
        filters.append(PerformanceMetric.end_date <= start_date_to)
    if min_profit is not None:
        filters.append(PerformanceMetric.net_profit >= min_profit)
    if max_drawdown is not None:
        filters.append(PerformanceMetric.max_drawdown <= max_drawdown)

    base_stmt = (
        select(PerformanceMetric, Strategy.name.label("strategy_name"), Instrument.symbol.label("instrument_symbol"))
        .options(load_only(PerformanceMetric.id, PerformanceMetric.user_id, PerformanceMetric.strategy_id, PerformanceMetric.instrument_id, PerformanceMetric.timeframe, PerformanceMetric.start_date, PerformanceMetric.end_date, PerformanceMetric.initial_capital, PerformanceMetric.final_capital, PerformanceMetric.net_profit, PerformanceMetric.max_drawdown, PerformanceMetric.sharpe_ratio, PerformanceMetric.sortino_ratio, PerformanceMetric.calmar_ratio, PerformanceMetric.win_rate, PerformanceMetric.total_trades, PerformanceMetric.winning_trades, PerformanceMetric.losing_trades, PerformanceMetric.profit_factor, PerformanceMetric.period, PerformanceMetric.return_pct, PerformanceMetric.avg_win, PerformanceMetric.avg_loss, PerformanceMetric.expectancy, PerformanceMetric.status, PerformanceMetric.created_at))
        .outerjoin(Strategy, Strategy.id == PerformanceMetric.strategy_id)
        .outerjoin(Instrument, Instrument.id == PerformanceMetric.instrument_id)
        .where(and_(*filters))
    )

    total = (
        await db.execute(
            select(func.count()).select_from(PerformanceMetric).where(and_(*filters))
        )
    ).scalar() or 0

    offset = (page - 1) * page_size
    rows = (
        await db.execute(base_stmt.order_by(desc(PerformanceMetric.created_at)).offset(offset).limit(page_size))
    ).all()

    backtest_ids = [str(row.id) for row, _, _ in rows]
    debit_map = await _get_backtest_debit_map(db, backtest_ids)

    items = []
    for row, strategy_name, instrument_symbol in rows:
        debit = debit_map.get(str(row.id), {})
        items.append(
            await _serialize_summary(
                db,
                row,
                strategy_name,
                instrument_symbol,
                credit_cost=debit.get("effective_credit_cost", debit.get("credit_cost")),
                debit_transaction_id=debit.get("debit_transaction_id"),
            )
        )

    total_pages = ceil(total / page_size) if total > 0 else 1
    return success_response(
        {
            "backtests": items,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_count": int(total),
                "total_pages": int(total_pages),
            },
        }
    )


@router.get("/{backtest_id}")
async def get_backtest_by_id(
    backtest_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    row = (await db.execute(select(PerformanceMetric).options(load_only(PerformanceMetric.id, PerformanceMetric.user_id, PerformanceMetric.strategy_id, PerformanceMetric.instrument_id, PerformanceMetric.timeframe, PerformanceMetric.start_date, PerformanceMetric.end_date, PerformanceMetric.initial_capital, PerformanceMetric.final_capital, PerformanceMetric.net_profit, PerformanceMetric.max_drawdown, PerformanceMetric.sharpe_ratio, PerformanceMetric.sortino_ratio, PerformanceMetric.calmar_ratio, PerformanceMetric.win_rate, PerformanceMetric.total_trades, PerformanceMetric.winning_trades, PerformanceMetric.losing_trades, PerformanceMetric.profit_factor, PerformanceMetric.period, PerformanceMetric.return_pct, PerformanceMetric.avg_win, PerformanceMetric.avg_loss, PerformanceMetric.expectancy, PerformanceMetric.status, PerformanceMetric.created_at)).where(PerformanceMetric.id == backtest_id))).scalars().first()
    if not row or str(row.user_id) != str(current_user["user_id"]):
        raise HTTPException(status_code=404, detail="Backtest not found")
    debit = (await _get_backtest_debit_map(db, [str(row.id)])).get(str(row.id), {})
    return success_response(
        await _serialize_summary(
            db,
            row,
            credit_cost=debit.get("effective_credit_cost", debit.get("credit_cost")),
            debit_transaction_id=debit.get("debit_transaction_id"),
        )
    )




def _jsonish(value):
    if value is None:
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return value
    return value




def _currency_symbol_for_code(code: str | None) -> str:
    normalized = (code or "").upper()
    if normalized == "USD":
        return "$"
    if normalized == "INR":
        return "₹"
    return code or "₹"


def _infer_currency_payload(*, instrument_symbol: str | None = None, asset_class: str | None = None, account_currency: str | None = None, currency_symbol: str | None = None, quantity_mode: str | None = None, instrument_spec: dict | None = None) -> dict:
    spec = instrument_spec if isinstance(instrument_spec, dict) else {}
    symbol_upper = (instrument_symbol or spec.get("symbol") or "").upper()
    asset_upper = (asset_class or spec.get("asset_class") or "").upper()
    account = account_currency or spec.get("account_currency")
    q_mode = quantity_mode or spec.get("quantity_mode")
    if not account:
        if symbol_upper in {"XAUUSD", "BTCUSD", "ETHUSD"} or asset_upper in {"METAL", "FOREX", "CRYPTO"}:
            account = "USD"
        else:
            account = "INR"
    if not currency_symbol:
        currency_symbol = spec.get("currency_symbol") or _currency_symbol_for_code(account)
    if not q_mode:
        if asset_upper in {"METAL", "FOREX", "CRYPTO"} or symbol_upper in {"XAUUSD", "BTCUSD", "ETHUSD"}:
            q_mode = "LOTS"
        elif asset_upper == "INDIAN_EQUITY":
            q_mode = "SHARES"
        else:
            q_mode = "SHARES"
    return {
        "account_currency": account,
        "currency_symbol": currency_symbol,
        "quantity_mode": q_mode,
        "asset_class": asset_class or spec.get("asset_class"),
        "is_legacy_currency": not bool(account_currency or spec.get("account_currency")),
    }

async def _professional_summary_overlay(db: AsyncSession, backtest_id: str, trades_data: list[dict]) -> dict:
    """Best-effort Phase 2F report metadata overlay for new and legacy runs."""
    overlay: dict = {}
    try:
        perf_cols = {meta["column_name"] for meta in await _table_columns_meta(db, "performance_metrics")}
        wanted = ["account_currency", "currency_symbol", "quantity_mode", "runtime_config_snapshot", "instrument_spec_snapshot", "professional_summary"]
        available = [col for col in wanted if col in perf_cols]
        if available:
            row = (await db.execute(text(f"SELECT {', '.join(available)} FROM performance_metrics WHERE id::text = :backtest_id LIMIT 1"), {"backtest_id": str(backtest_id)})).mappings().first()
            if row:
                for key in ["account_currency", "currency_symbol", "quantity_mode"]:
                    if key in row and row.get(key) is not None:
                        overlay[key] = row.get(key)
                runtime_config = _jsonish(row.get("runtime_config_snapshot")) if "runtime_config_snapshot" in row else None
                instrument_spec = _jsonish(row.get("instrument_spec_snapshot")) if "instrument_spec_snapshot" in row else None
                professional_summary = _jsonish(row.get("professional_summary")) if "professional_summary" in row else None
                if runtime_config is not None:
                    overlay["runtime_config_snapshot"] = runtime_config
                if instrument_spec is not None:
                    overlay["instrument_spec_snapshot"] = instrument_spec
                if isinstance(professional_summary, dict):
                    for key in ["asset_class", "avg_actual_risk", "avg_lot_size", "avg_quantity", "gross_profit", "gross_loss"]:
                        if professional_summary.get(key) is not None:
                            overlay[key] = professional_summary.get(key)
    except Exception as exc:
        logger.warning("Unable to load professional summary overlay for backtest %s: %s", backtest_id, exc)

    first_trade = next((trade for trade in trades_data if isinstance(trade, dict)), None)
    if first_trade:
        for key in ["account_currency", "currency_symbol", "asset_class", "quantity_mode", "sl_mode", "position_size_mode"]:
            if not overlay.get(key) and first_trade.get(key) is not None:
                overlay[key] = first_trade.get(key)
        if not overlay.get("runtime_config_snapshot") and first_trade.get("runtime_config_snapshot") is not None:
            overlay["runtime_config_snapshot"] = first_trade.get("runtime_config_snapshot")
        if not overlay.get("instrument_spec_snapshot") and first_trade.get("instrument_spec_snapshot") is not None:
            overlay["instrument_spec_snapshot"] = first_trade.get("instrument_spec_snapshot")

    runtime_config = overlay.get("runtime_config_snapshot")
    if isinstance(runtime_config, dict):
        risk = runtime_config.get("risk") or {}
        sl_tp = runtime_config.get("sl_tp") or {}
        if isinstance(risk, dict):
            if not overlay.get("position_size_mode") and risk.get("position_size_mode"):
                overlay["position_size_mode"] = risk.get("position_size_mode")
            if overlay.get("risk_percent") is None and risk.get("risk_percent") is not None:
                overlay["risk_percent"] = risk.get("risk_percent")
        if isinstance(sl_tp, dict):
            if overlay.get("rr_ratio") is None and sl_tp.get("rr_ratio") is not None:
                overlay["rr_ratio"] = sl_tp.get("rr_ratio")
            if not overlay.get("sl_mode") and sl_tp.get("sl_mode"):
                overlay["sl_mode"] = sl_tp.get("sl_mode")

    instrument_spec = overlay.get("instrument_spec_snapshot")
    if isinstance(instrument_spec, dict):
        for src, dst in [("asset_class", "asset_class"), ("account_currency", "account_currency"), ("currency_symbol", "currency_symbol"), ("quantity_mode", "quantity_mode")]:
            if not overlay.get(dst) and instrument_spec.get(src) is not None:
                overlay[dst] = instrument_spec.get(src)

    if trades_data:
        actual_risks = [_to_float(t.get("actual_risk_amount"), None) for t in trades_data if _to_float(t.get("actual_risk_amount"), None) is not None]
        lot_sizes = [_to_float(t.get("lot_size"), None) for t in trades_data if _to_float(t.get("lot_size"), None) is not None]
        quantities = [_to_float(t.get("quantity"), None) for t in trades_data if _to_float(t.get("quantity"), None) is not None]
        pnls = [_to_float(t.get("pnl"), 0.0) for t in trades_data]
        if actual_risks and overlay.get("avg_actual_risk") is None:
            overlay["avg_actual_risk"] = sum(actual_risks) / len(actual_risks)
        if lot_sizes and overlay.get("avg_lot_size") is None:
            overlay["avg_lot_size"] = sum(lot_sizes) / len(lot_sizes)
        if quantities and overlay.get("avg_quantity") is None:
            overlay["avg_quantity"] = sum(quantities) / len(quantities)
        if overlay.get("gross_profit") is None:
            overlay["gross_profit"] = sum(v for v in pnls if v > 0)
        if overlay.get("gross_loss") is None:
            overlay["gross_loss"] = sum(v for v in pnls if v < 0)

    inferred = _infer_currency_payload(
        instrument_symbol=overlay.get("instrument_symbol"),
        asset_class=overlay.get("asset_class"),
        account_currency=overlay.get("account_currency"),
        currency_symbol=overlay.get("currency_symbol"),
        quantity_mode=overlay.get("quantity_mode"),
        instrument_spec=instrument_spec if isinstance(instrument_spec, dict) else None,
    )
    for key, value in inferred.items():
        if overlay.get(key) is None or overlay.get(key) == "Legacy":
            overlay[key] = value
    return overlay

@router.get("/{backtest_id}/detail")
async def get_backtest_detail(
    backtest_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    row = (await db.execute(select(PerformanceMetric).options(load_only(PerformanceMetric.id, PerformanceMetric.user_id, PerformanceMetric.strategy_id, PerformanceMetric.instrument_id, PerformanceMetric.timeframe, PerformanceMetric.start_date, PerformanceMetric.end_date, PerformanceMetric.initial_capital, PerformanceMetric.final_capital, PerformanceMetric.net_profit, PerformanceMetric.max_drawdown, PerformanceMetric.sharpe_ratio, PerformanceMetric.sortino_ratio, PerformanceMetric.calmar_ratio, PerformanceMetric.win_rate, PerformanceMetric.total_trades, PerformanceMetric.winning_trades, PerformanceMetric.losing_trades, PerformanceMetric.profit_factor, PerformanceMetric.period, PerformanceMetric.return_pct, PerformanceMetric.avg_win, PerformanceMetric.avg_loss, PerformanceMetric.expectancy, PerformanceMetric.status, PerformanceMetric.created_at)).where(PerformanceMetric.id == backtest_id))).scalars().first()
    if not row or str(row.user_id) != str(current_user["user_id"]):
        raise HTTPException(status_code=404, detail="Backtest not found")

    trades_data = []
    equity_data = []
    pnl_data = []

    try:
        trade_columns = [meta["column_name"] for meta in await _table_columns_meta(db, "trades")]
        if trade_columns:
            wanted = [
                "id", "entry_time", "exit_time", "side", "quantity", "lot_size", "entry_price", "exit_price", "pnl", "exit_type", "exit_reason",
                "account_currency", "currency_symbol", "asset_class", "quantity_mode",
                "stop_loss", "target", "risk_points", "risk_ticks", "risk_pips", "reward_points", "reward_ticks",
                "rr_ratio", "risk_amount", "actual_risk_amount", "reward_amount", "expected_reward_amount", "r_multiple",
                "sl_mode", "position_size_mode", "runtime_config_snapshot", "instrument_spec_snapshot", "lifecycle_events", "signal_reason",
            ]
            select_columns = [column for column in wanted if column in trade_columns]
            select_sql = ", ".join(select_columns)
            rows = (
                await db.execute(
                    text(f"SELECT {select_sql} FROM trades WHERE backtest_id::text = :backtest_id ORDER BY entry_time ASC"),
                    {"backtest_id": str(backtest_id)},
                )
            ).mappings().all()
            for row_map in rows:
                row_dict = dict(row_map)
                risk_values = _risk_values_from_row(row_dict)
                trades_data.append(
                    {
                        "id": str(row_dict.get("id")) if row_dict.get("id") is not None else None,
                        "entry_time": row_dict.get("entry_time").isoformat() if row_dict.get("entry_time") else None,
                        "exit_time": row_dict.get("exit_time").isoformat() if row_dict.get("exit_time") else None,
                        "side": row_dict.get("side"),
                        "quantity": _to_float(row_dict.get("quantity"), None),
                        "lot_size": _to_float(row_dict.get("lot_size"), None),
                        "entry_price": _to_float(row_dict.get("entry_price")),
                        "exit_price": _to_float(row_dict.get("exit_price")),
                        "pnl": _to_float(row_dict.get("pnl")),
                        "exit_type": row_dict.get("exit_type") or row_dict.get("exit_reason"),
                        "exit_reason": row_dict.get("exit_reason") or row_dict.get("exit_type"),
                        "account_currency": row_dict.get("account_currency"),
                        "currency_symbol": row_dict.get("currency_symbol"),
                        "asset_class": row_dict.get("asset_class"),
                        "quantity_mode": row_dict.get("quantity_mode"),
                        "actual_risk_amount": _to_float(row_dict.get("actual_risk_amount"), None),
                        "risk_ticks": _to_float(row_dict.get("risk_ticks"), None),
                        "risk_pips": _to_float(row_dict.get("risk_pips"), None),
                        "reward_ticks": _to_float(row_dict.get("reward_ticks"), None),
                        "expected_reward_amount": _to_float(row_dict.get("expected_reward_amount"), None),
                        "sl_mode": row_dict.get("sl_mode"),
                        "position_size_mode": row_dict.get("position_size_mode"),
                        "runtime_config_snapshot": _jsonish(row_dict.get("runtime_config_snapshot")),
                        "instrument_spec_snapshot": _jsonish(row_dict.get("instrument_spec_snapshot")),
                        "lifecycle_events": _jsonish(row_dict.get("lifecycle_events")) or [],
                        **risk_values,
                    }
                )
    except Exception as exc:
        logger.warning("Unable to load trades for backtest %s: %s", backtest_id, exc)

    if not trades_data:
        try:
            if await _column_exists(db, "performance_metrics", "trade_details"):
                fallback_row = (
                    await db.execute(
                        text("SELECT trade_details FROM performance_metrics WHERE id::text = :backtest_id LIMIT 1"),
                        {"backtest_id": str(backtest_id)},
                    )
                ).mappings().first()
                if fallback_row:
                    trades_data = _normalise_trade_detail_rows(fallback_row.get("trade_details"))
        except Exception as exc:
            logger.warning("Unable to load fallback trade_details for backtest %s: %s", backtest_id, exc)

    try:
        equity_rows = (
            await db.execute(
                select(EquityCurve)
                .where(cast(EquityCurve.backtest_id, String) == str(backtest_id))
                .order_by(EquityCurve.timestamp.asc())
            )
        ).scalars().all()
        equity_data = [
            {
                "timestamp": e.timestamp.isoformat() if e.timestamp else None,
                "equity": _to_float(e.equity),
            }
            for e in equity_rows
        ]
    except Exception as exc:
        logger.warning("Unable to load equity for backtest %s: %s", backtest_id, exc)

    try:
        pnl_rows = (
            await db.execute(
                select(PnLCalendar)
                .where(cast(PnLCalendar.backtest_id, String) == str(backtest_id))
                .order_by(PnLCalendar.date.asc())
            )
        ).scalars().all()
        pnl_data = [
            {
                "date": p.date.isoformat() if p.date else None,
                "pnl": _to_float(p.pnl),
            }
            for p in pnl_rows
        ]
    except Exception as exc:
        logger.warning("Unable to load pnl calendar for backtest %s: %s", backtest_id, exc)

    debit = (await _get_backtest_debit_map(db, [str(row.id)])).get(str(row.id), {})
    summary = await _serialize_summary(
        db,
        row,
        credit_cost=debit.get("effective_credit_cost", debit.get("credit_cost")),
        debit_transaction_id=debit.get("debit_transaction_id"),
    )
    summary.update(await _professional_summary_overlay(db, str(backtest_id), trades_data))
    currency_payload = _infer_currency_payload(
        instrument_symbol=summary.get("instrument_symbol"),
        asset_class=summary.get("asset_class"),
        account_currency=summary.get("account_currency"),
        currency_symbol=summary.get("currency_symbol"),
        quantity_mode=summary.get("quantity_mode"),
        instrument_spec=summary.get("instrument_spec_snapshot") if isinstance(summary.get("instrument_spec_snapshot"), dict) else None,
    )
    summary.update({k: v for k, v in currency_payload.items() if summary.get(k) in (None, "Legacy")})
    for trade in trades_data:
        trade_payload = _infer_currency_payload(
            instrument_symbol=summary.get("instrument_symbol"),
            asset_class=trade.get("asset_class") or summary.get("asset_class"),
            account_currency=trade.get("account_currency") or summary.get("account_currency"),
            currency_symbol=trade.get("currency_symbol") or summary.get("currency_symbol"),
            quantity_mode=trade.get("quantity_mode") or summary.get("quantity_mode"),
            instrument_spec=trade.get("instrument_spec_snapshot") if isinstance(trade.get("instrument_spec_snapshot"), dict) else None,
        )
        for key, value in trade_payload.items():
            if trade.get(key) in (None, "Legacy"):
                trade[key] = value
    return success_response(
        {
            "summary": summary,
            "trades": trades_data,
            "equity_curve": equity_data,
            "pnl_calendar": pnl_data,
        }
    )
